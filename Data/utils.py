import math
import random

import numpy as np
import openpyxl


class Utils:
    def __init__(self, vehicles):
        self.MSE_data = list()
        self.it = list()
        self.vehicles = vehicles

    def distance_part(self, g, n_data, part, part_ant, distances, array_part, dfirst=False):

        """
        Calculates the distance between (x,y)^t and (x,y)^(t-1).
        """
        if dfirst:
            part_ant[0, 2 * n_data] = part[0]
            part_ant[0, 2 * n_data + 1] = part[1]
            # if n_data == 1:
            #    part_ant[0, 0] = part[0]
            #   part_ant[0, 1] = part[1]
            # if n_data == 2:
            #   part_ant[0, 2] = part[0]
            #  part_ant[0, 3] = part[1]
            # if n_data == 3:
            #   part_ant[0, 4] = part[0]
            #  part_ant[0, 5] = part[1]
            # if n_data == 4:
            #   part_ant[0, 6] = part[0]
            #  part_ant[0, 7] = part[1]
        else:
            array_part[0, 2 * n_data] = part[0]
            array_part[0, 2 * n_data + 1] = part[1]
            distances[n_data] = math.sqrt(
                (array_part[0, 2 * n_data] - part_ant[g, 2 * n_data]) ** 2 + (
                        array_part[0, 2 * n_data + 1] - part_ant[g, 2 * n_data + 1])
                ** 2) + distances[n_data]

            # if n_data == 1:
            #   array_part[0, 0] = part[0]
            #  array_part[0, 1] = part[1]
            # distances[0] = math.sqrt(
            #    (array_part[0, 0] - part_ant[g, 0]) ** 2 + (array_part[0, 1] - part_ant[g, 1])
            #   ** 2) + distances[0]
            # elif n_data == 2:
            #   array_part[0, 2] = part[0]
            #  array_part[0, 3] = part[1]
            # distances[1] = math.sqrt(
            #    (array_part[0, 2] - part_ant[g, 2]) ** 2 + (array_part[0, 3] - part_ant[g, 3])
            #   ** 2) + distances[1]
            # elif n_data == 3:
            #   array_part[0, 4] = part[0]
            #  array_part[0, 5] = part[1]
            # distances[2] = math.sqrt(
            #    (array_part[0, 4] - part_ant[g, 4]) ** 2 + (array_part[0, 5] - part_ant[g, 5])
            #   ** 2) + distances[2]
            # elif n_data == 4:
            #   array_part[0, 6] = part[0]
            #  array_part[0, 7] = part[1]
            # distances[3] = math.sqrt(
            #    (array_part[0, 6] - part_ant[g, 6]) ** 2 + (array_part[0, 7] - part_ant[g, 7])
            #   ** 2) + distances[3]
            if n_data == self.vehicles - 1:
                part_ant = np.append(part_ant, array_part, axis=0)

        return part_ant, distances

    def mse(self, g, y_data, mu_data):

        """
        Calculates the mean square error (MSE) between the collected data and the estimated data.
        """

        total_suma = 0
        y_array = np.array(y_data)
        mu_array = np.array(mu_data)
        for i in range(len(mu_array)):
            total_suma = ((float(y_array[i]) - float(mu_array[i])) ** 2) + total_suma
        MSE = total_suma / y_data.shape[0]
        self.MSE_data.append(MSE)
        self.it.append(g)
        return self.MSE_data, self.it

    def savexlsx(self, sigma_data, mu_data, seed, file):

        """
        Save the data in excel documents.
        """

        for i in range(len(mu_data)):
            mu_data[i] = float(mu_data[i])

        wb1 = openpyxl.Workbook()
        hoja1 = wb1.active
        hoja1.append(self.MSE_data)
        wb1.save('Test/' + file + '/ALLCONError' + str(seed[0]) + '.xlsx')

        wb2 = openpyxl.Workbook()
        hoja2 = wb2.active
        hoja2.append(sigma_data)
        wb2.save('Test/' + file + '/ALLCONSigma' + str(seed[0]) + '.xlsx')

        wb3 = openpyxl.Workbook()
        hoja3 = wb3.active
        hoja3.append(mu_data)
        wb3.save('Test/' + file + '/ALLCONMu' + str(seed[0]) + '.xlsx')

        wb4 = openpyxl.Workbook()
        hoja4 = wb4.active
        hoja4.append(list(self.distances))
        wb4.save('Test/' + file + '/ALLCONDistance' + str(seed[0]) + '.xlsx')

        wb5 = openpyxl.Workbook()
        hoja5 = wb5.active
        hoja5.append(self.it)
        wb5.save('Test/' + file + '/ALLCONData' + str(seed[0]) + '.xlsx')


def obtain_prefabricated_vehicles(vehicle_number, subfleet_number):
    """
        Función que retorna una tupla de vehículos y sus sensores dependiendo del numero de subflotas que se requieren

    @param numero_de_subflotas: 1, 2, 3 o 4 subflotas
    @return: (list[str], list[list[str]]): tupla que corresponde a los vehículos (que pueden renombrarse a cualquier
                                           cosa) y a los sensores que tiene dicho vehículo según el numero_de_subflotas
                                           seleccionado.
    -1: antes
    -2: sensores >= 4
    -3: sensores >= 2
    """
    if vehicle_number == -1:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v7', 'v8'],
                 [['s1', 's6'], ['s3', 's1'], ['s3'], ['s4'], ['s1'], ['s4', 's3'], ['s6']]),
                (['v1', 'v2', 'v3', 'v5', 'v6', 'v7'],
                 [['s1', 's6'], ['s8', 's5'], ['s5'], ['s5', 's1'], ['s6'],
                  ['s8', 's6']]),
                (['v3', 'v6', 'v7', 'v8'],
                 [['s8', 's4'], ['s7', 's8'], ['s7', 's4', 's2'], ['s8', 's4', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v6', 'v7'],
                 [['s1'], ['s4', 's1'], ['s2', 's4'], ['s4'], ['s2', 's1'], ['s7', 's1']]),
                (['v1', 'v3', 'v5', 'v6', 'v7'],
                 [['s1'], ['s3'], ['s3', 's2', 's1'], ['s7', 's2', 's8'], ['s8', 's1']]),
                (['v2', 'v3', 'v4', 'v5', 'v7', 'v8'],
                 [['s2'], ['s8', 's1'], ['s4'], ['s1', 's4'], ['s1', 's8'], ['s2', 's8']]),
                (['v1', 'v2', 'v6', 'v7', 'v8'],
                 [['s1', 's3'], ['s5', 's8'], ['s2', 's6'], ['s2', 's5'], ['s8', 's1']]),
                (['v2', 'v4', 'v5', 'v7', 'v8'],
                 [['s2', 's6', 's7'], ['s1', 's6'], ['s7'], ['s7', 's5'], ['s8', 's2']]),
                (['v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s4', 's5'], ['s4'], ['s5'], ['s8'], ['s1'], ['s8', 's1'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v5', 'v7', 'v8'],
                 [['s7', 's4'], ['s1'], ['s4'], ['s6', 's4'], ['s7', 's6'], ['s1', 's6']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v5', 'v9', 'v10', 'v12', 'v13'],
                 [['s9'], ['s4'], ['s8'], ['s9', 's1'], ['s1'], ['s4'], ['s9'], ['s8', 's4']]),
                (['v1', 'v3', 'v4', 'v6', 'v8', 'v10'],
                 [['s9'], ['s3', 's6', 's9'], ['s9', 's3'], ['s6'], ['s8'], ['s8', 's2']]),
                (['v2', 'v3', 'v4', 'v5'],
                 [['s3', 's4', 's2'], ['s1', 's6', 's7'], ['s4', 's2'], ['s5', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's4'], ['s2'], ['s3', 's2', 's5'], ['s4', 's6'], ['s5', 's7']]),
                (['v1', 'v2', 'v5', 'v7', 'v8'],
                 [['s1'], ['s4', 's7', 's1'], ['s7', 's4'], ['s2', 's8'], ['s8', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8'], ['s7'], ['s4'], ['s4', 's8'], ['s4'], ['s4', 's8'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3'], ['s8'], ['s3'], ['s4', 's7'], ['s4', 's3'], ['s7'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's8'], ['s4', 's8'], ['s1', 's4'], ['s6'], ['s6', 's5'], ['s4']]),
                (['v1', 'v2', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1'], ['s2', 's6'], ['s1'], ['s3', 's4'], ['s2'], ['s3', 's1'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v6', 'v7', 'v8'],
                 [['s1'], ['s3'], ['s3', 's6', 's4'], ['s4'], ['s6'], ['s1'], ['s8', 's1']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v3', 'v4', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s3', 's1'], ['s6', 's10'], ['s3'], ['s7'], ['s5', 's3'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v9'],
                 [['s1'], ['s1', 's5'], ['s10'], ['s5'], ['s6', 's8'], ['s5', 's10'], ['s9']]),
                (['v1', 'v3', 'v4', 'v5', 'v6', 'v7', 'v9', 'v10'],
                 [['s2'], ['s10'], ['s7', 's10'], ['s5'], ['s10'], ['s10', 's8'], ['s7'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1'], ['s5'], ['s5'], ['s4'], ['s9'], ['s9'], ['s5'], ['s9', 's5'], ['s9']]),
                (['v1', 'v2', 'v4', 'v5', 'v6', 'v8', 'v9', 'v10'],
                 [['s1', 's3'], ['s10'], ['s1'], ['s5'], ['s5'], ['s8', 's1'], ['s10'], ['s10']]),
                (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v9', 'v10'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s7'], ['s9', 's3'], ['s7', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v9', 'v10'],
                 [['s5'], ['s10'], ['s5', 's8'], ['s2'], ['s10'], ['s10'], ['s8'], ['s8'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2'], ['s2', 's10'], ['s6'], ['s4'], ['s1'], ['s6'], ['s4'], ['s10']]),
                (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v9', 'v10'],
                 [['s1'], ['s2'], ['s3'], ['s1', 's6'], ['s4'], ['s4'], ['s4', 's3'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7'], ['s8'], ['s9'], ['s9'], ['s8'], ['s8'], ['s8'], ['s8'], ['s7']])])

    elif vehicle_number == 523:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
            [['s1'], ['s1'], ['s2'], ['s2', 's3'], ['s3']])

    elif vehicle_number == 524:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
                    [['s1', 's4'], ['s1', 's4'], ['s2'], ['s2', 's3'], ['s3']])

    elif vehicle_number == 525:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
                    [['s1', 's4'], ['s1', 's4'], ['s2', 's5'], ['s2', 's3'], ['s3', 's5']])

    elif vehicle_number == 526:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
                    [['s1', 's4', 's6'], ['s1', 's4'], ['s2', 's5', 's6'], ['s2', 's3'], ['s3', 's5']])

    elif vehicle_number == 527:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
                    [['s1', 's4', 's6'], ['s1', 's4', 's7'], ['s2', 's5', 's6'], ['s2', 's3', 's7'], ['s3', 's5']])

    elif vehicle_number == 532:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
            [['s1'], ['s1'], ['s2'], ['s2', 's1'], ['s2']])

    elif vehicle_number == 533:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
            [['s1'], ['s1', 's3'], ['s2', 's3'], ['s2', 's1'], ['s2', 's3']])

    elif vehicle_number == 534:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
            [['s1', 's4'], ['s1', 's3', 's4'], ['s2', 's3', 's4'], ['s2', 's1'], ['s2', 's3']])

    elif vehicle_number == 535:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
            [['s1', 's4', 's5'], ['s1', 's3', 's4'], ['s2', 's3', 's4'], ['s2', 's1', 's5'], ['s2', 's3', 's5']])

    elif vehicle_number == 542:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
            [['s1'], ['s1', 's2'], ['s1', 's2'], ['s2', 's1'], ['s2']])

    elif vehicle_number == 543:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5'],
            [['s1', 's3'], ['s1', 's2', 's3'], ['s1', 's2'], ['s2', 's1', 's3'], ['s2', 's3']])

    elif vehicle_number == 623:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s3'], ['s3']])

    elif vehicle_number == 624:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4'], ['s1'], ['s2', 's4'], ['s2'], ['s3'], ['s3']])

    elif vehicle_number == 625:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4'], ['s1', 's5'], ['s2', 's4'], ['s2'], ['s3', 's5'], ['s3']])

    elif vehicle_number == 626:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4'], ['s1', 's5'], ['s2', 's4'], ['s2', 's6'], ['s3', 's5'], ['s3', 's6']])

    elif vehicle_number == 627:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4', 's7'], ['s1', 's5'], ['s2', 's4'], ['s2', 's6'], ['s3', 's5', 's7'], ['s3', 's6']])

    elif vehicle_number == 628:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4', 's7'], ['s1', 's5', 's8'], ['s2', 's4'], ['s2', 's6', 's8'], ['s3', 's5', 's7'], ['s3', 's6']])

    elif vehicle_number == 629:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4', 's7'], ['s1', 's5', 's8'], ['s2', 's4', 's9'], ['s2', 's6', 's8'], ['s3', 's5', 's7'], ['s3', 's6', 's9']])

    elif vehicle_number == 632:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s1'], ['s2']])

    elif vehicle_number == 633:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1'], ['s1', 's3'], ['s2', 's3'], ['s2'], ['s1', 's2'], ['s3']])

    elif vehicle_number == 634:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4'], ['s1', 's3'], ['s2', 's3'], ['s2', 's4'], ['s1', 's2'], ['s3', 's4']])

    elif vehicle_number == 635:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4', 's5'], ['s1', 's3'], ['s2', 's3', 's5'], ['s2', 's4', 's5'], ['s1', 's2'], ['s3', 's4']])

    elif vehicle_number == 636:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4', 's5'], ['s1', 's3', 's6'], ['s2', 's3', 's5'], ['s2', 's4', 's5'], ['s1', 's2', 's6'], ['s3', 's4', 's6']])

    elif vehicle_number == 642:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1'], ['s1'], ['s2', 's1'], ['s2'], ['s1', 's2'], ['s2']])

    elif vehicle_number == 643:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's3'], ['s1', 's3'], ['s2', 's1'], ['s2', 's3'], ['s1', 's2'], ['s2', 's3']])

    elif vehicle_number == 644:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's4', 's3'], ['s1', 's3', 's4'], ['s2', 's1'], ['s2', 's3', 's4'], ['s1', 's2', 's4'], ['s2', 's3']])

    elif vehicle_number == 652:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's2'], ['s1'], ['s2', 's1'], ['s2'], ['s1', 's2'], ['s2', 's1']])

    elif vehicle_number == 653:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            [['s1', 's2', 's3'], ['s1', 's3'], ['s2', 's1', 's3'], ['s2', 's3'], ['s1', 's2', 's3'], ['s2', 's1']])

    elif vehicle_number == 724:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s3'], ['s3', 's4'], ['s4']])

    elif vehicle_number == 725:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1'], ['s1', 's5'], ['s2', 's5'], ['s2'], ['s3'], ['s3', 's4'], ['s4']])

    elif vehicle_number == 726:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's6'], ['s1', 's5'], ['s2', 's5'], ['s2'], ['s3', 's6'], ['s3', 's4'], ['s4']])

    elif vehicle_number == 727:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's6'], ['s1', 's5'], ['s2', 's5'], ['s2', 's7'], ['s3', 's6'], ['s3', 's4'], ['s4', 's7']])

    elif vehicle_number == 728:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's6', 's8'], ['s1', 's5'], ['s2', 's5'], ['s2', 's7', 's8'], ['s3', 's6'], ['s3', 's4'], ['s4', 's7']])

    elif vehicle_number == 729:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's6', 's8'], ['s1', 's5', 's9'], ['s2', 's5'], ['s2', 's7', 's8'], ['s3', 's6'], ['s3', 's4'], ['s4', 's7', 's9']])

    elif vehicle_number == 7210:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's6', 's8'], ['s1', 's5', 's9'], ['s2', 's5', 's10'], ['s2', 's7', 's8'], ['s3', 's6', 's10'], ['s3', 's4'], ['s4', 's7', 's9']])

    elif vehicle_number == 733:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1'], ['s1'], ['s2'], ['s2', 's3'], ['s3'], ['s3', 's1'], ['s2']])

    elif vehicle_number == 734:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1'], ['s1', 's4'], ['s2', 's4'], ['s2', 's3'], ['s3', 's4'], ['s3', 's1'], ['s2']])

    elif vehicle_number == 735:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's5'], ['s1', 's4'], ['s2', 's4', 's5'], ['s2', 's3'], ['s3', 's4'], ['s3', 's1'], ['s2', 's5']])

    elif vehicle_number == 736:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's5', 's6'], ['s1', 's4'], ['s2', 's4', 's5'], ['s2', 's3', 's6'], ['s3', 's4'], ['s3', 's1'], ['s2', 's5', 's6']])

    elif vehicle_number == 737:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's5', 's6'], ['s1', 's4', 's7'], ['s2', 's4', 's5'], ['s2', 's3', 's6'], ['s3', 's4', 's7'], ['s3', 's1', 's7'], ['s2', 's5', 's6']])

    elif vehicle_number == 742:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1'], ['s1'], ['s2'], ['s2', 's1'], ['s2'], ['s1'], ['s2']])

    elif vehicle_number == 743:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's3'], ['s1', 's3'], ['s2', 's3'], ['s2', 's1'], ['s2', 's3'], ['s1'], ['s2']])

    elif vehicle_number == 744:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's3'], ['s1', 's3'], ['s2', 's3'], ['s2', 's1', 's4'], ['s2', 's3', 's4'], ['s1', 's4'], ['s2', 's4']])

    elif vehicle_number == 745:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's3', 's5'], ['s1', 's3', 's5'], ['s2', 's3', 's5'], ['s2', 's1', 's4'], ['s2', 's3', 's4'], ['s1', 's4', 's5'], ['s2', 's4', 's5']])

    elif vehicle_number == 752:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1'], ['s1'], ['s2', 's1'], ['s2', 's1'], ['s2'], ['s1', 's2'], ['s2']])

    elif vehicle_number == 753:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's3'], ['s1', 's3'], ['s2', 's1'], ['s2', 's1'], ['s2', 's3'], ['s1', 's2', 's3'], ['s2', 's3']])

    elif vehicle_number == 754:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's3', 's4'], ['s1', 's3', 's4'], ['s2', 's1', 's4'], ['s2', 's1'], ['s2', 's3', 's4'], ['s1', 's2', 's3'], ['s2', 's3', 's4']])

    elif vehicle_number == 762:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1'], ['s1', 's2'], ['s2', 's1'], ['s2', 's1'], ['s2', 's1'], ['s1', 's2'], ['s2']])

    elif vehicle_number == 763:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
            [['s1', 's3'], ['s1', 's2', 's3'], ['s2', 's1', 's3'], ['s2', 's1', 's3'], ['s2', 's1', 's3'], ['s1', 's2'], ['s2', 's3']])

    elif vehicle_number == 824:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s3'], ['s3'], ['s4'], ['s4']])

    elif vehicle_number == 825:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's5'], ['s1'], ['s2', 's5'], ['s2'], ['s3'], ['s3'], ['s4'], ['s4']])

    elif vehicle_number == 826:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's5'], ['s1', 's6'], ['s2', 's5'], ['s2'], ['s3', 's6'], ['s3'], ['s4'], ['s4']])

    elif vehicle_number == 827:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's5'], ['s1', 's6'], ['s2', 's5'], ['s2'], ['s3', 's6'], ['s3', 's7'], ['s4', 's7'], ['s4']])

    elif vehicle_number == 828:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's5'], ['s1', 's6'], ['s2', 's5'], ['s2', 's8'], ['s3', 's6'], ['s3', 's7'], ['s4', 's7'], ['s4', 's8']])

    elif vehicle_number == 829:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's5', 's9'], ['s1', 's6'], ['s2', 's5'], ['s2', 's8', 's9'], ['s3', 's6'], ['s3', 's7'], ['s4', 's7'], ['s4', 's8']])

    elif vehicle_number == 8210:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's5', 's9'], ['s1', 's6', 's10'], ['s2', 's5'], ['s2', 's8', 's9'], ['s3', 's6'], ['s3', 's7'], ['s4', 's7', 's10'], ['s4', 's8']])

    elif vehicle_number == 833:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1'], ['s1'], ['s2'], ['s2'], ['s3'], ['s3'], ['s1', 's3'], ['s2']])

    elif vehicle_number == 834:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's4'], ['s1'], ['s2', 's4'], ['s2'], ['s3', 's4'], ['s3'], ['s1', 's3'], ['s2']])

    elif vehicle_number == 835:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's4'], ['s1', 's5'], ['s2', 's4'], ['s2', 's5'], ['s3', 's4'], ['s3', 's5'], ['s1', 's3'], ['s2']])

    elif vehicle_number == 836:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's4'], ['s1', 's5', 's6'], ['s2', 's4'], ['s2', 's5'], ['s3', 's4', 's6'], ['s3', 's5'], ['s1', 's3'], ['s2', 's6']])

    elif vehicle_number == 837:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's4', 's7'], ['s1', 's5', 's6'], ['s2', 's4'], ['s2', 's5', 's7'], ['s3', 's4', 's6'], ['s3', 's5'], ['s1', 's3', 's7'], ['s2', 's6']])

    elif vehicle_number == 838:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                    [['s1', 's4', 's7'], ['s1', 's5', 's6'], ['s2', 's4', 's8'], ['s2', 's5', 's7'], ['s3', 's4', 's6'], ['s3', 's5', 's8'], ['s1', 's3', 's7'], ['s2', 's6', 's8']])

    elif vehicle_number == 842:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s2'], ['s2'], ['s1'], ['s1']])

    elif vehicle_number == 843:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3'], ['s3'], ['s2', 's3'], ['s2', 's1'], ['s2'], ['s2', 's3'], ['s1'], ['s1']])

    elif vehicle_number == 844:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3'], ['s3', 's4'], ['s2', 's3'], ['s2', 's1'], ['s2', 's4'], ['s2', 's3'], ['s1', 's4'], ['s1', 's4']])

    elif vehicle_number == 845:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3', 's5'], ['s3', 's4'], ['s2', 's3', 's5'], ['s2', 's1'], ['s2', 's4', 's5'], ['s2', 's3'], ['s1', 's4', 's5'], ['s1', 's4']])

    elif vehicle_number == 846:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3', 's5'], ['s3', 's4', 's6'], ['s2', 's3', 's5'], ['s2', 's1', 's6'], ['s2', 's4', 's5'], ['s2', 's3', 's6'], ['s1', 's4', 's5'], ['s1', 's4', 's6']])

    elif vehicle_number == 852:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1'], ['s1'], ['s2', 's1'], ['s2'], ['s2'], ['s2'], ['s1', 's2'], ['s1']])

    elif vehicle_number == 853:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3'], ['s1', 's3'], ['s2', 's1'], ['s2'], ['s2', 's3'], ['s2', 's3'], ['s1', 's2'], ['s1', 's3']])

    elif vehicle_number == 854:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3', 's4'], ['s1', 's3'], ['s2', 's1', 's4'], ['s2', 's4'], ['s2', 's3'], ['s2', 's3', 's4'], ['s1', 's2'], ['s1', 's3', 's4']])

    elif vehicle_number == 862:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1'], ['s1', 's2'], ['s2', 's1'], ['s2'], ['s2'], ['s2', 's1'], ['s1', 's2'], ['s1']])

    elif vehicle_number == 863:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3'], ['s1', 's2'], ['s2', 's1', 's3'], ['s2', 's3'], ['s2', 's3'], ['s2', 's1'], ['s1', 's2', 's3'], ['s1', 's3']])

    elif vehicle_number == 864:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
            [['s1', 's3', 's4'], ['s1', 's2', 's4'], ['s2', 's1', 's3'], ['s2', 's3', 's4'], ['s2', 's3', 's4'], ['s2', 's1', 's4'], ['s1', 's2', 's3'], ['s1', 's3', 's4']])

    elif vehicle_number == 925:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s3'], ['s3'], ['s4'], ['s4', 's5'], ['s5']])

    elif vehicle_number == 926:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's6'], ['s1'], ['s2', 's6'], ['s2'], ['s3'], ['s3'], ['s4'], ['s4', 's5'], ['s5']])

    elif vehicle_number == 927:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's6'], ['s1', 's7'], ['s2', 's6'], ['s2'], ['s3', 's7'], ['s3'], ['s4'], ['s4', 's5'], ['s5']])

    elif vehicle_number == 928:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's6'], ['s1', 's7'], ['s2', 's6'], ['s2', 's8'], ['s3', 's7'], ['s3'], ['s4'], ['s4', 's5'], ['s5', 's8']])

    elif vehicle_number == 929:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's6'], ['s1', 's7'], ['s2', 's6'], ['s2', 's8'], ['s3', 's7'], ['s3', 's9'], ['s4', 's9'], ['s4', 's5'], ['s5', 's8']])

    elif vehicle_number == 9210:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's6', 's10'], ['s1', 's7'], ['s2', 's6'], ['s2', 's8'], ['s3', 's7'], ['s3', 's9', 's10'], ['s4', 's9'], ['s4', 's5'], ['s5', 's8']])

    elif vehicle_number == 933:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s3'], ['s3'], ['s1'], ['s2'], ['s3']])

    elif vehicle_number == 934:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1'], ['s1', 's4'], ['s2'], ['s2', 's4'], ['s3'], ['s3', 's4'], ['s1'], ['s2'], ['s3']])

    elif vehicle_number == 935:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's5'], ['s1', 's4'], ['s2', 's5'], ['s2', 's4'], ['s3'], ['s3', 's4'], ['s1'], ['s2'], ['s3', 's5']])

    elif vehicle_number == 936:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's5'], ['s1', 's4'], ['s2', 's5'], ['s2', 's4'], ['s3', 's6'], ['s3', 's4'], ['s1', 's6'], ['s2', 's6'], ['s3', 's5']])

    elif vehicle_number == 937:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's5', 's7'], ['s1', 's4'], ['s2', 's5'], ['s2', 's4'], ['s3', 's6', 's7'], ['s3', 's4'], ['s1', 's6', 's7'], ['s2', 's6'], ['s3', 's5']])

    elif vehicle_number == 938:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's5', 's7'], ['s1', 's4', 's8'], ['s2', 's5'], ['s2', 's4', 's8'], ['s3', 's6', 's7'], ['s3', 's4'], ['s1', 's6', 's7'], ['s2', 's6', 's8'], ['s3', 's5']])

    elif vehicle_number == 938:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's5', 's7'], ['s1', 's4', 's8'], ['s2', 's5', 's9'], ['s2', 's4', 's8'], ['s3', 's6', 's7'],
             ['s3', 's4', 's9'], ['s1', 's6', 's7'], ['s2', 's6', 's8'], ['s3', 's5', 's9']])

    elif vehicle_number == 943:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1'], ['s1', 's3'], ['s2'], ['s2'], ['s3', 's2'], ['s3'], ['s3'], ['s1'], ['s2', 's1']])

    elif vehicle_number == 944:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                    [['s1', 's4'], ['s1', 's3'], ['s2', 's4'], ['s2'], ['s3', 's2'], ['s3'], ['s3', 's4'], ['s1', 's4'], ['s2', 's1']])

    elif vehicle_number == 945:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                    [['s1', 's4'], ['s1', 's3', 's5'], ['s2', 's4', 's5'], ['s2', 's5'], ['s3', 's2'], ['s3', 's5'], ['s3', 's4'], ['s1', 's4'], ['s2', 's1']])

    elif vehicle_number == 946:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                    [['s1', 's4', 's6'], ['s1', 's3', 's5'], ['s2', 's4', 's5'], ['s2', 's5', 's6'], ['s3', 's2'], ['s3', 's5', 's6'], ['s3', 's4'], ['s1', 's4', 's6'], ['s2', 's1']])

    elif vehicle_number == 952:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s2'], ['s2'], ['s1'], ['s1'], ['s2', 's1']])

    elif vehicle_number == 953:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's3'], ['s1'], ['s2', 's3'], ['s2', 's3'], ['s2'], ['s3'], ['s1', 's3'], ['s1', 's2'], ['s2', 's1']])

    elif vehicle_number == 954:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's3'], ['s1', 's4'], ['s2', 's3'], ['s2', 's3'], ['s2', 's4'], ['s3', 's4'], ['s1', 's3', 's4'], ['s1', 's2'], ['s2', 's1', 's4']])

    elif vehicle_number == 955:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's3', 's5'], ['s1', 's4', 's5'], ['s2', 's3'], ['s2', 's3', 's5'], ['s2', 's4'], ['s3', 's4', 's5'], ['s1', 's3', 's4'], ['s1', 's2', 's5'], ['s2', 's1', 's4']])

    elif vehicle_number == 962:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s2'], ['s2', 's1'], ['s1'], ['s1', 's2'], ['s2', 's1']])

    elif vehicle_number == 963:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's3'], ['s1', 's3'], ['s2', 's3'], ['s2', 's3'], ['s2', 's3'], ['s2', 's1'], ['s1', 's3'], ['s1', 's2'], ['s2', 's1']])

    elif vehicle_number == 964:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
            [['s1', 's3'], ['s1', 's3', 's4'], ['s2', 's3', 's4'], ['s2', 's3', 's4'], ['s2', 's3', 's4'], ['s2', 's1'], ['s1', 's3', 's4'], ['s1', 's2', 's4'], ['s2', 's1']])

    elif vehicle_number == 1025:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s3'], ['s3'], ['s4'], ['s4'], ['s5'], ['s5']])

    elif vehicle_number == 1026:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1', 's6'], ['s2'], ['s2'], ['s3', 's6'], ['s3'], ['s4'], ['s4'], ['s5'], ['s5']])

    elif vehicle_number == 1027:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1', 's6'], ['s2'], ['s2'], ['s3', 's6'], ['s3', 's7'], ['s4'], ['s4'], ['s5', 's7'], ['s5']])

    elif vehicle_number == 1028:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1', 's6'], ['s2'], ['s2', 's8'], ['s3', 's6'], ['s3', 's7'], ['s4'], ['s4', 's8'], ['s5', 's7'], ['s5']])

    elif vehicle_number == 1029:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's9'], ['s1', 's6'], ['s2'], ['s2', 's8'], ['s3', 's6'], ['s3', 's7'], ['s4', 's9'], ['s4', 's8'], ['s5', 's7'], ['s5']])

    elif vehicle_number == 10210:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's9'], ['s1', 's6'], ['s2', 's10'], ['s2', 's8'], ['s3', 's6'], ['s3', 's7'], ['s4', 's9'], ['s4', 's8'], ['s5', 's7'], ['s5', '10']])

    elif vehicle_number == 1034:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1'], ['s2'], ['s2', 's4'], ['s3'], ['s3'], ['s4'], ['s4', 's3'], ['s1'], ['s2']])

    elif vehicle_number == 1035:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1', 's5'], ['s2', 's5'], ['s2', 's4'], ['s3'], ['s3', 's5'], ['s4'], ['s4', 's3'], ['s1'], ['s2']])

    elif vehicle_number == 1036:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's6'], ['s1', 's5'], ['s2', 's5'], ['s2', 's4'], ['s3', 's6'], ['s3', 's5'], ['s4', 's6'], ['s4', 's3'], ['s1'], ['s2']])

    elif vehicle_number == 1037:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's6'], ['s1', 's5'], ['s2', 's5', 's7'], ['s2', 's4'], ['s3', 's6'], ['s3', 's5'], ['s4', 's6'], ['s4', 's3'], ['s1', 's7'], ['s7', 's2']])

    elif vehicle_number == 1038:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's6', 's8'], ['s1', 's5'], ['s2', 's5', 's7'], ['s2', 's4', 's8'], ['s3', 's6'], ['s3', 's5', 's8'], ['s4', 's6'], ['s4', 's3'], ['s1', 's7'], ['s7', 's2']])

    elif vehicle_number == 1039:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's6', 's8'], ['s1', 's5', 's9'], ['s2', 's5', 's7'], ['s2', 's4', 's8'], ['s3', 's6'],
             ['s3', 's5', 's8'], ['s4', 's6', 's9'], ['s4', 's3'], ['s1', 's7'], ['s7', 's2', 's9']])

    elif vehicle_number == 10310:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's6', 's8'], ['s1', 's5', 's9'], ['s2', 's5', 's7'], ['s2', 's4', 's8'], ['s3', 's6', 's10'],
             ['s3', 's5', 's8'], ['s4', 's6', 's9'], ['s4', 's3', 's10'], ['s1', 's7', 's10'], ['s7', 's2', 's9']])

    elif vehicle_number == 1043:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1'], ['s2'], ['s2', 's1'], ['s3'], ['s3'], ['s3'], ['s2', 's3'], ['s1'], ['s2']])

    elif vehicle_number == 1044:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1', 's4'], ['s2', 's4'], ['s2', 's1'], ['s3', 's4'], ['s3'], ['s3', 's1'], ['s2', 's3'], ['s4'], ['s2']])

    elif vehicle_number == 1045:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's5'], ['s1', 's4'], ['s2', 's4'], ['s2', 's1'], ['s3', 's4'], ['s3', 's5'], ['s3', 's1'], ['s2', 's3'], ['s4', 's5'], ['s2', 's5']])

    elif vehicle_number == 1046:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's5', 's6'], ['s1', 's4'], ['s2', 's4', 's6'], ['s2', 's1'], ['s3', 's4', 's6'], ['s3', 's5', 's6'], ['s3', 's1'], ['s2', 's3'], ['s4', 's5'], ['s2', 's5']])

    elif vehicle_number == 1047:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's5', 's6'], ['s1', 's4', 's7'], ['s2', 's4', 's6'], ['s2', 's1', 's7'],
             ['s3', 's4', 's6'], ['s3', 's5', 's6'], ['s3', 's1', 's7'], ['s2', 's3'], ['s4', 's5', 's7'], ['s2', 's5']])

    elif vehicle_number == 1052:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1'], ['s2'], ['s2'], ['s1'], ['s2'], ['s1'], ['s2'], ['s1'], ['s2']])

    elif vehicle_number == 1053:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's3'], ['s1'], ['s2', 's3'], ['s3'], ['s1', 's2'], ['s2', 's3'], ['s1', 's3'], ['s2'], ['s1'], ['s2']])

    elif vehicle_number == 1054:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's3'], ['s1', 's4'], ['s2', 's3'], ['s3', 's4'], ['s1', 's2'], ['s2', 's3'], ['s1', 's3'], ['s2', 's4'], ['s1', 's4'], ['s2', 's4']])

    elif vehicle_number == 1055:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's3', 's5'], ['s1', 's4'], ['s2', 's3', 's5'], ['s3', 's4'], ['s1', 's2', 's5'], ['s2', 's3'], ['s1', 's3'], ['s2', 's4', 's5'], ['s1', 's4', 's5'], ['s2', 's4']])

    elif vehicle_number == 1056:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's3', 's5'], ['s1', 's4', 's6'], ['s2', 's3', 's5'], ['s3', 's4', 's6'], ['s1', 's2', 's5'],
             ['s2', 's3', 's6'], ['s1', 's3', 's6'], ['s2', 's4', 's5'], ['s1', 's4', 's5'], ['s2', 's4', 's6']])

    elif vehicle_number == 1062:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1'], ['s2', 's1'], ['s2'], ['s1', 's2'], ['s2'], ['s1'], ['s2'], ['s1'], ['s2']])

    elif vehicle_number == 1063:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1'], ['s1', 's3'], ['s2', 's1'], ['s2', 's3'], ['s1', 's2'], ['s2', 's3'], ['s3'], ['s2', 's1'], ['s1', 's3'], ['s2', 's3']])

    elif vehicle_number == 1064:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's4'], ['s1', 's3'], ['s2', 's1', 's4'], ['s2', 's3'], ['s1', 's2'], ['s2', 's3'], ['s3', 's4'], ['s2', 's1', 's4'], ['s1', 's3', 's4'], ['s2', 's3', 's4']])

    elif vehicle_number == 1065:
        if subfleet_number == 1:
            return (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
            [['s1', 's4', 's5'], ['s1', 's3', 's5'], ['s2', 's1', 's4'], ['s2', 's3', 's5'], ['s1', 's2', 's5'],
             ['s2', 's3', 's5'], ['s3', 's4', 's5'], ['s2', 's1', 's4'], ['s1', 's3', 's4'], ['s2', 's3', 's4']])

    elif vehicle_number == -2:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s1', 's2'], ['s1', 's2'], ['s2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's3'], ['s1', 's3'], ['s1', 's2', 's3'], ['s1', 's2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s1', 's2'], ['s1', 's2'], ['s1'], ['s2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2'], ['s1', 's2'], ['s1', 's3'], ['s2', 's3'], ['s1', 's3'], ['s2', 's3']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                #  [['s1', 's2', 's3'], ['s1', 's2'], ['s1'], ['s1', 's2', 's3'], ['s3'], ['s2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s1', 's2', 's3'], ['s1', 's3'], ['s1', 's2'], ['s3'], ['s2', 's3'], ['s2']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                #  [['s1', 's2'], ['s1', 's3', 's4'], ['s1', 's4'], ['s2', 's3'], ['s2', 's4'],
                #   ['s3', 's4'], ['s1', 's2', 's3']]),
                (['p1', 'p2', 'p3', 'p4', 'p5', 'p6', 'p7', 'p8'],
                 [['s1', 's2'], ['s1', 's3'], ['s1', 's4'], ['s1', 's2'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4'], ['s3', 's4']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                #  [['s1', 's2'], ['s1', 's3', 's2'], ['s4'], ['s1'], ['s2', 's3'], ['s2', 's4'],
                #   ['s3', 's4'], ['s1', 's3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's3'], ['s1', 's4'], ['s1', 's4'], ['s2', 's3'], ['s2'], ['s2', 's3'],
                  ['s4'], ['s3', 's4']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                #  [['s1', 's2'], ['s1', 's3', 's5'], ['s1', 's4'], ['s1', 's4', 's5'], ['s2', 's3'],
                #   ['s2', 's5'], ['s2', 's3'], ['s4', 's5'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2', 's5'], ['s1'], ['s1', 's3', 's5'], ['s1', 's3', 's4'], ['s4', 's5'], ['s2'],
                  ['s2', 's3', 's4'], ['s4'], ['s2', 's3', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5'], ['s2', 's3'], ['s3', 's5'], ['s1', 's4'], ['s1', 's3'], ['s1', 's4'],
                  ['s2', 's4'], ['s2', 's5'], ['s2', 's5'], ['s3', 's4']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                #  [['s1', 's5', 's6'], ['s2', 's6'], ['s3', 's6'], ['s1', 's4', 's5'], ['s1', 's3'],
                #   ['s1', 's4'], ['s2', 's4'], ['s2', 's5', 's3'], ['s2', 's5'], ['s3', 's4', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2', 's5'], ['s2', 's3', 's4'], ['s3', 's4', 's5'], ['s1'], ['s3'],
                  ['s1', 's3', 's4'], ['s2'], ['s1', 's2', 's5'], ['s5'], ['s4']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                #  [['s1', 's5', 's6'], ['s6'], ['s3', 's6'], ['s1', 's4', 's5'], ['s3'],
                #   ['s1', 's2', 's4'], ['s1', 's2', 's4'], ['s2', 's5', 's3'], ['s2', 's5'], ['s3', 's4', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5', 's6'], ['s2', 's6', 's7'], ['s3', 's6'], ['s1', 's4', 's5'], ['s1', 's3'],
                  ['s1', 's4', 's7'], ['s2', 's4', 's7'], ['s2', 's5', 's3'], ['s2', 's5', 's7'], ['s3', 's4', 's6']]),

                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s1', 's2'], ['s1', 's2'], ['s2', 's1'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's3', 's1'], ['s1', 's3', 's2'], ['s1', 's2', 's3'], ['s1', 's2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s1', 's2'], ['s1', 's2'], ['s1', 's2'], ['s2', 's1'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2', 's3'], ['s1', 's2'], ['s1', 's3', 's2'], ['s2', 's3'], ['s1', 's3'], ['s2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2', 's3'], ['s1', 's2', 's3'], ['s1'], ['s1', 's2', 's3'], ['s3', 's1'], ['s2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s1', 's2', 's3'], ['s1', 's3', 's4'], ['s1', 's2', 's4'], ['s3', 's4'], ['s2', 's3'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's2'], ['s1', 's3', 's4'], ['s1', 's4'], ['s2', 's3', 's1'], ['s2', 's4'],
                  ['s3', 's4', 's2'], ['s1', 's2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2', 's3'], ['s1', 's3'], ['s1', 's4', 's2'], ['s1', 's2', 's4'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4', 's1'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s1', 's3', 's2'], ['s4', 's2'], ['s1'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4'], ['s1', 's3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's3', 's4'], ['s1', 's4'], ['s1', 's4'], ['s2', 's3'], ['s2', 's3'], ['s2', 's3'],
                  ['s4', 's3'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's3', 's5'], ['s1', 's4'], ['s1', 's4', 's5'], ['s2', 's3', 's5'],
                  ['s2', 's5'], ['s2', 's3', 's1'], ['s4', 's5', 's3'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2', 's5'], ['s1', 's3'], ['s1', 's3', 's5'], ['s1', 's3', 's4'], ['s2', 's4', 's5'],
                  ['s2', 's3', 's4'], ['s4', 's1'], ['s2', 's3', 's5'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5'], ['s2', 's3', 's5'], ['s3', 's5'], ['s1', 's4'], ['s1', 's3', 's4'], ['s1', 's4'],
                  ['s2', 's4'], ['s2', 's5'], ['s2', 's5'], ['s3', 's4', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5', 's6'], ['s2', 's6'], ['s3', 's6'], ['s1', 's4', 's5'], ['s1', 's3', 's2'],
                  ['s1', 's4', 's5'], ['s2', 's4'], ['s2', 's5', 's3'], ['s2', 's5', 's6'], ['s3', 's4', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2', 's5'], ['s2', 's3', 's4'], ['s3', 's4', 's5'], ['s1', 's4'], ['s3', 's5'],
                  ['s1', 's3', 's4'], ['s2', 's1'], ['s1', 's2', 's5'], ['s5', 's4'], ['s4', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5', 's6'], ['s6'], ['s3', 's6', 's1'], ['s1', 's4', 's5'], ['s3'],
                  ['s1', 's2', 's4'], ['s1', 's2', 's4'], ['s2', 's5', 's3'], ['s2', 's5', 's3'],
                  ['s3', 's4', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5', 's6'], ['s2', 's6', 's7'], ['s3', 's6', 's2'], ['s1', 's4', 's5'], ['s1', 's3', 's7'],
                  ['s1', 's4', 's7'], ['s2', 's4', 's7'], ['s2', 's5', 's3'], ['s2', 's5', 's7'],
                  ['s3', 's4', 's6']])
                ])

    elif vehicle_number == -3:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s1', 's2'], ['s1', 's2'], ['s2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's3'], ['s1', 's3'], ['s1', 's2', 's3'], ['s1', 's2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s1', 's2'], ['s1'], ['s2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's3'], ['s1'], ['s1', 's2'], ['s1', 's3']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s1', 's2'], ['s1', 's2'], ['s1'], ['s2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2'], ['s1', 's2'], ['s1', 's3'], ['s2', 's3'], ['s1', 's3'], ['s2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2', 's3'], ['s1', 's2'], ['s1'], ['s1', 's2', 's3'], ['s3'], ['s2', 's3']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s2'], ['s1', 's2'], ['s1'], ['s2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2'], ['s1', 's2'], ['s3'], ['s2', 's3'], ['s3'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2'], ['s1', 's2'], ['s1'], ['s1', 's3'], ['s3'], ['s2', 's3']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s1', 's2', 's3'], ['s1', 's3'], ['s1', 's2'], ['s3'], ['s2', 's3'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's2'], ['s1', 's3', 's4'], ['s1', 's4'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4'], ['s1', 's2', 's3']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s1', 's2'], ['s1', 's3'], ['s1'], ['s3'], ['s2', 's3'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s2'], ['s1', 's4'], ['s1', 's4'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4'], ['s1', 's3']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s1', 's3'], ['s1', 's4'], ['s1', 's2'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s1', 's3', 's2'], ['s4'], ['s1'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4'], ['s1', 's3', 's4']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s1', 's3'], ['s1'], ['s1'], ['s2', 's3'], ['s2', 's4'],
                  ['s3', 's4'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s1', 's2'], ['s4'], ['s1'], ['s2', 's3'], ['s4'],
                  ['s3', 's4'], ['s1', 's4']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's3'], ['s1', 's4'], ['s1', 's4'], ['s2', 's3'], ['s2'], ['s2', 's3'],
                  ['s4'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's3', 's5'], ['s1', 's4'], ['s1', 's4', 's5'], ['s2', 's3'],
                  ['s2', 's5'], ['s2', 's3'], ['s4', 's5'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2', 's5'], ['s1'], ['s1', 's3', 's5'], ['s1', 's3', 's4'], ['s4', 's5'], ['s2'],
                  ['s2', 's3', 's4'], ['s4'], ['s2', 's3', 's5']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s3'], ['s1', 's4'], ['s4'], ['s2'], ['s2'], ['s2', 's3'],
                  ['s4'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's5'], ['s1', 's4'], ['s1', 's5'], ['s2', 's3'],
                  ['s2', 's5'], ['s2', 's3'], ['s4'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1'], ['s1', 's5'], ['s1', 's3'], ['s4', 's5'], ['s2'],
                  ['s3', 's4'], ['s4'], ['s2', 's5']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5'], ['s2', 's3'], ['s5'], ['s1'], ['s1', 's3'], ['s1', 's4'],
                  ['s2', 's4'], ['s2', 's5'], ['s2'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5', 's6'], ['s2', 's6'], ['s3', 's6'], ['s1', 's4', 's5'], ['s1', 's3'],
                  ['s1', 's4'], ['s2', 's4'], ['s2', 's5', 's3'], ['s2', 's5'], ['s3', 's4', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2', 's5'], ['s2', 's4'], ['s3', 's5'], ['s1'], ['s3'],
                  ['s1', 's3', 's4'], ['s2'], ['s1', 's2',], ['s5'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5', 's6'], ['s6'], ['s3', 's6'], ['s1', 's4', 's5'], ['s3'],
                  ['s1', 's2', 's4'], ['s1', 's2', 's4'], ['s2', 's5', 's3'], ['s2', 's5'], ['s3', 's4', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5', 's6'], ['s2', 's6', 's7'], ['s3'], ['s1', 's4', 's5'], ['s1', 's3'],
                  ['s1', 's4', 's7'], ['s2', 's4', 's7'], ['s2', 's5'], ['s2', 's5'],
                  ['s4', 's6']]),

                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s1', 's2'], ['s1', 's2'], ['s1'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's3', 's1'], ['s1', 's2'], ['s1', 's2', 's3'], ['s1', 's2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s1', 's2'], ['s1', 's2'], ['s1'], ['s2', 's1'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2', 's3'], ['s1', 's2'], ['s1', 's2'], ['s2', 's3'], ['s1', 's3'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's3'], ['s1', 's3'], ['s1'], ['s1', 's2'], ['s3', 's1'], ['s2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s1', 's2'], ['s1', 's3'], ['s1', 's2', 's4'], ['s3', 's4'],
                  ['s2', 's3'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's2'], ['s1', 's4'], ['s1', 's4'], ['s2', 's3', 's1'], ['s2', 's4'],
                  ['s3', 's2'], ['s1', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s1', 's3'], ['s1', 's2'], ['s1', 's4'], ['s2', 's3'],
                  ['s2', 's4'],
                  ['s3', 's4', 's1'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s1', 's2'], ['s4', 's2'], ['s1'], ['s2'], ['s2', 's4'],
                  ['s3', 's4'], ['s1', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's4'], ['s1'], ['s1', 's4'], ['s2', 's3'], ['s2', 's3'],
                  ['s2'],
                  ['s4', 's3'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's5'], ['s1', 's4'], ['s1', 's4', 's5'], ['s2', 's3', 's5'],
                  ['s2', 's5'], ['s2', 's1'], ['s4', 's3'], ['s3', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2', 's5'], ['s1', 's3'], ['s1', 's5'], ['s1', 's3', 's4'], ['s2', 's4', 's5'],
                  ['s2', 's3', 's4'], ['s4'], ['s2', 's3', 's5'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5'], ['s2', 's3'], ['s3', 's5'], ['s1', 's4'], ['s1', 's4'], ['s1', 's4'],
                  ['s2', 's4'], ['s2', 's5'], ['s2'], ['s3', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's6'], ['s2'], ['s3', 's6'], ['s1', 's5'], ['s1', 's3', 's2'],
                  ['s1', 's5'], ['s2', 's4'], ['s2', 's3'], ['s2', 's5', 's6'], ['s3', 's4', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2'], ['s3', 's4'], ['s3', 's5'], ['s1', 's4'], ['s3', 's5'],
                  ['s1', 's3', 's4'], ['s2', 's1'], [ 's2', 's5'], ['s5', 's4'], ['s4', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5', 's6'], ['s6'], ['s6', 's1'], ['s1', 's4', 's5'], ['s3'],
                  ['s1', 's4'], ['s1', 's4'], ['s2', 's5', 's3'], ['s2', 's5', 's3'],
                  ['s3', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5'], ['s2', 's6', 's7'], ['s3', 's6', 's2'], ['s1', 's4', 's5'],
                  ['s1', 's3'],
                  ['s1', 's4', 's7'], ['s4', 's7'], ['s2', 's3'], ['s2', 's5', 's7'],
                  ['s3', 's4', 's6']])
            ])

    elif vehicle_number == -4:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's6'], ['s3', 's1'], ['s3'], ['s4'], ['s1'], ['s4', 's3'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's6'], ['s8', 's5'], ['s5'], ['s5', 's1'], ['s6'],
                  ['s8', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s8', 's4'], ['s7', 's8'], ['s7', 's4', 's2'], ['s7', 's2'], ['s8', 's4', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s4', 's1'], ['s2', 's4'], ['s4', 's7'], ['s2', 's1'], ['s7', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's7'], ['s3'], ['s3', 's2', 's1'], ['s7', 's2', 's8'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s2'], ['s8', 's1'], ['s4'], ['s1', 's4'], ['s1', 's8'], ['s2', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's3'], ['s5', 's8', 's3'], ['s2', 's6'], ['s2', 's5', 's6'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s2', 's6', 's7'], ['s1', 's6'], ['s7', 's1'], ['s7', 's5'], ['s8', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5'], ['s4'], ['s5'], ['s8'], ['s1'], ['s8', 's1'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s7', 's4'], ['s1'], ['s4'], ['s6', 's4'], ['s7', 's6'], ['s1', 's6']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s9'], ['s4'], ['s8'], ['s9', 's1'], ['s1'], ['s4'], ['s9'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s9'], ['s3', 's6', 's9'], ['s9', 's3'], ['s6'], ['s8', 's2'], ['s8', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s3', 's4', 's2'], ['s1', 's6', 's7'], ['s1', 's6', 's7'], ['s4', 's2'], ['s3', 's5', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's4'], ['s2', 's3', 's7'], ['s3', 's2', 's5'], ['s4', 's6', 's1'], ['s5', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1'], ['s4', 's7', 's1'], ['s7', 's4'], ['s2', 's8', 's9'], ['s8', 's9', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8'], ['s7'], ['s4'], ['s4', 's8'], ['s4'], ['s4', 's8'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3'], ['s8'], ['s3'], ['s4', 's7'], ['s4', 's3'], ['s7'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's8'], ['s4', 's8'], ['s1', 's4'], ['s6', 's5'], ['s6', 's5'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s2', 's6'], ['s1'], ['s3', 's4'], ['s2', 's6'], ['s3', 's1'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s3'], ['s3', 's6', 's4'], ['s4'], ['s6'], ['s1', 's8'], ['s8', 's1']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's1'], ['s7', 's6'], ['s5', 's3', 's7'], ['s3', 's10'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1'], ['s1', 's5'], ['s10', 's9'], ['s5'], ['s6', 's8'], ['s6', 's8'], ['s5', 's10'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s2'], ['s10'], ['s7', 's10'], ['s2'], ['s10'], ['s10', 's8'], ['s7', 's8'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1'], ['s5'], ['s5'], ['s1'], ['s9'], ['s9'], ['s5'], ['s9', 's5'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's3'], ['s10'], ['s1'], ['s5'], ['s5'], ['s3', 's1'], ['s10'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s7'], ['s9', 's3'], ['s7', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5'], ['s10'], ['s5', 's8'], ['s5'], ['s10'], ['s10'], ['s8'], ['s8'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s2', 's10'], ['s6'], ['s4'], ['s1'], ['s10'], ['s4'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1'], ['s2'], ['s3'], ['s1', 's6'], ['s4', 's6'], ['s4'], ['s4', 's3'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7'], ['s8'], ['s9'], ['s9'], ['s8'], ['s8'], ['s8'], ['s8'], ['s7']])])

    elif vehicle_number == -5:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's6'], ['s3', 's1'], ['s3'], ['s4'], ['s1', 's3'], ['s4', 's3'], ['s6', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's6'], ['s8', 's5'], ['s5', 's6'], ['s5', 's1'], ['s6', 's1'],
                  ['s8', 's6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's4'], ['s7', 's8'], ['s7', 's4', 's2'], ['s7', 's2', 's4'], ['s8', 's4', 's2'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s4', 's1'], ['s2', 's4'], ['s4', 's7'], ['s2', 's1'], ['s7', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s3', 's1'], ['s3', 's2', 's1'], ['s3', 's2', 's8'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's8', 's6'], ['s2', 's6', 's1'], ['s2', 's1', 's6'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s2', 's6', 's7'], ['s1', 's6'], ['s7', 's1'], ['s7', 's2'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5'], ['s4'], ['s5', 's8'], ['s8'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s9', 's8'], ['s4', 's9'], ['s8'], ['s9', 's1'], ['s1', 's4'], ['s4'], ['s9'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's7'], ['s7','s4'], ['s4'], ['s4', 's8'], ['s4'], ['s4', 's8'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3'], ['s8', 's7'], ['s3'], ['s4', 's7'], ['s4', 's3'], ['s7'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s2', 's6'], ['s1', 's4'], ['s3', 's4', 's2'], ['s2', 's6'], ['s3', 's1'], ['s2', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s3'], ['s3', 's6', 's4'], ['s4', 's1'], ['s6'], ['s1', 's3'], ['s8', 's4']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1'], ['s5'], ['s5', 's1'], ['s1'], ['s9'], ['s9'], ['s5'], ['s9', 's5'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5'], ['s10'], ['s5', 's8'], ['s5'], ['s10'], ['s10'], ['s8'], ['s8'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's9'], ['s8'], ['s9'], ['s9'], ['s8'], ['s8'], ['s8'], ['s8'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1'], ['s5'], ['s5'], ['s1'], ['s9'], ['s9'], ['s5'], ['s9'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5'], ['s10'], ['s5', 's8'], ['s5'], ['s10'], ['s10'], ['s8'], ['s8', 's7'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's9'], ['s8'], ['s9'], ['s9'], ['s8'], ['s7'], ['s8'], ['s8'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's9'], ['s8'], ['s9'], ['s9'], ['s8'], ['s7'], ['s8', 's10'], ['s8', 's10'], ['s7']])])

    elif vehicle_number == -6:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s3', 's1'], ['s3'], ['s4'], ['s1', 's3'], ['s4', 's3'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's6'], ['s8'], ['s6'], ['s1'], ['s6', 's1'],
                  ['s8', 's6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's4'], ['s7', 's8'], ['s7', 's4'], ['s7', 's4'], ['s8', 's4'], ['s8']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                #  [['s1'], ['s4', 's1'], ['s2', 's4'], ['s4'], ['s2', 's1'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's8'], ['s3', 's1', 's8'], ['s3', 's1'], ['s3', 's8'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's8', 's6'], ['s2', 's6', 's1'], ['s2', 's1', 's6'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s2', 's7'], ['s1', 's2'], ['s7', 's1'], ['s7', 's2'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5'], ['s4', 's5'], ['s5', 's8'], ['s8'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s9', 's8'], ['s4', 's9'], ['s8'], ['s9', 's1'], ['s1', 's4'], ['s4'], ['s9'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's7'], ['s4'], ['s4'], ['s4', 's8'], ['s4'], ['s4', 's8'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3'], ['s7'], ['s3', 's4'], ['s4', 's7'], ['s4', 's3'], ['s7'], ['s7'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s2', 's6'], ['s1', 's4'], ['s4', 's2'], ['s2', 's6'], ['s4', 's1'], ['s2', 's1']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1'], ['s5'], ['s5', 's1'], ['s1'], ['s9'], ['s9'], ['s5'], ['s9', 's5'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5'], ['s10'], ['s8'], ['s5', 's8'], ['s10'], ['s10'], ['s8'], ['s8'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's9'], ['s9'], ['s9'], ['s9'], ['s8'], ['s8'], ['s8'], ['s8'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1'], ['s5'], ['s5'], ['s1'], ['s9'], ['s9'], ['s5'], ['s9'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5'], ['s10'], ['s5'], ['s5'], ['s10'], ['s10'], ['s5'], ['s7'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's9'], ['s8'], ['s9'], ['s9'], ['s8'], ['s7'], ['s8'], ['s8'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's9'], ['s8'], ['s9'], ['s9'], ['s8'], ['s7'], ['s8', 's10'], ['s8', 's10'], ['s7']])])

    elif vehicle_number == -7:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s1'], ['s3'], ['s4'], ['s3'], ['s2'], ['s2']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 #[['s1'], ['s8'], ['s6'], ['s1'], ['s6'],
                  #['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's4'], ['s7', 's8'], ['s4'], ['s7'], ['s1'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1'], ['s1', 's8'], ['s3'], ['s3'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's8'], ['s6', 's1'], ['s3', 's6'], ['s8', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s2', 's7'], ['s1', 's2', 's5'], ['s7', 's1', 's3'], ['s3', 's4'], ['s5', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5'], ['s8'], ['s1', 's8'], ['s2'], ['s1', 's2'], ['s5'], ['s4']]),

                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 #[['s9'], ['s9'], ['s8'], ['s1'], ['s1'], ['s4'], ['s8'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's7'], ['s1'], ['s4'], ['s2'], ['s4', 's2'], ['s1'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3'], ['s7'], ['s3', 's4'], ['s4', 's7'], ['s1', 's2'], ['s1', 's5'], ['s2'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s2', 's6'], ['s1', 's8'], ['s4'], ['s2', 's7'], ['s6', 's8'], ['s7']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1'], ['s5'], ['s5', 's1'], ['s2'], ['s6'], ['s3'], ['s3', 's2'], ['s4', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5'], ['s10'], ['s1'], ['s5', 's8'], ['s1'], ['s2'], ['s8'], ['s2'], ['s10']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 #[['s7'], ['s7'], ['s9'], ['s9'], ['s1'], ['s8'], ['s8'], ['s1'], ['s2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's8'], ['s5'], ['s4'], ['s4', 's7'], ['s3', 's8'], ['s3', 's7'], ['s2'], ['s2', 's6'],
                  ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5'], ['s10'], ['s5'], ['s1'], ['s10'], ['s1'], ['s2', 's4'], ['s3'], ['s4'], ['s2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's9'], ['s8'], ['s9', 's5'], ['s1', 's5'], ['s8'], ['s2', 's4'], ['s2'], ['s3', 's4'],
                  ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1'], ['s7', 's9'], ['s8', 's1'], ['s9', 's2'], ['s2', 's3'], ['s3', 's4'], ['s4', 's5'],
                  ['s5', 's10'], ['s6', 's10'], ['s6', 's8']])])

    elif vehicle_number == -8:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's3'], ['s1'], ['s3'], ['s2', 's1'], ['s3'], ['s2'], ['s2']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 #[['s1'], ['s1'], ['s6'], ['s1'], ['s6'],
                  #['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's7'], ['s7', 's8'], ['s8'], ['s7', 's1'], ['s1'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1'], ['s1', 's8'], ['s1'], ['s8'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's8'], ['s8', 's1'], ['s8', 's2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s2', 's7'], ['s1', 's2', 's5'], ['s7', 's1', 's2'], ['s5', 's7'], ['s1', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5'], ['s8'], ['s1', 's8'], ['s4', 's1'], ['s1', 's5'], ['s5', 's8'], ['s4']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s9'], ['s9'], ['s9'], ['s1'], ['s1', 's4'], ['s4'], ['s1'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's7'], ['s1'], ['s4'], ['s7', 's1'], ['s4', 's8'], ['s1', 's4'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3', 's1'], ['s7', 's2'], ['s3', 's4'], ['s4', 's7'], ['s1', 's2'], ['s1', 's7'],
                  ['s2', 's4'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s2', 's6'], ['s1', 's8', 's2'], ['s4', 's6'], ['s2', 's4'], ['s6', 's8'],
                  ['s8', 's1']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1'], ['s5'], ['s5', 's1'], ['s2'], ['s2', 's4'], ['s5'], ['s2'], ['s4', 's1'], ['s4']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 #[['s5'], ['s3'], ['s1'], ['s5'], ['s1'], ['s5'], ['s3'], ['s3'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7'], ['s7'], ['s1'], ['s1', 's2'], ['s8'], ['s8'], ['s1'], ['s8', 's2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's8'], ['s5'], ['s4'], ['s4', 's7'], ['s5', 's8'], ['s5', 's7'], ['s1'], ['s1', 's7'],
                  ['s4', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5'], ['s10'], ['s5'], ['s1'], ['s10', 's5'], ['s1', 's4'], ['s2', 's4'], ['s2', 's1'], ['s4'],
                  ['s2', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1'], ['s7', 's9'], ['s8', 's1'], ['s9', 's5'], ['s1', 's5'], ['s8', 's5'], ['s7', 's4'],
                  ['s9'], ['s8', 's4'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1'], ['s7', 's9', 's4'], ['s8', 's1'], ['s9', 's2'], ['s2', 's3'], ['s3', 's4'],
                  ['s4', 's8'],
                  ['s7', 's3'], ['s2', 's9'], ['s1', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1', 's6'], ['s7', 's9', 's4'], ['s8', 's1'], ['s9', 's2', 's6'], ['s2', 's3'], ['s3', 's4'],
                  ['s4', 's8', 's6'],
                  ['s7', 's3'], ['s2', 's9'], ['s1', 's8']])])

    elif vehicle_number == -9:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s1'], ['s1'], ['s2', 's1'], ['s2'], ['s2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s1'], ['s6', 's1'], ['s1', 's6'], ['s6'],
                  ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's7'], ['s7', 's8', 's1'], ['s8', 's7'], ['s7', 's1'], ['s1', 's8'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1'], ['s1', 's8'], ['s1', 's8'], ['s8', 's1'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2', 's8'], ['s2', 's8'], ['s8', 's1'], ['s8', 's2', 's1'], ['s1', 's2']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5'],
                #  [['s2', 's7', 's1'], ['s1', 's2', 's5'], ['s7', 's1', 's2'], ['s5', 's7', 's2'], ['s1', 's5', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5', 's8'], ['s8', 's4'], ['s1', 's8'], ['s4', 's1', 's5'], ['s1', 's5'], ['s5', 's8'], ['s4', 's1']]),

                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 #[['s9'], ['s9'], ['s9'], ['s9'], ['s4'], ['s4'], ['s4'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's7'], ['s1'], ['s8'], ['s7', 's1'], ['s1', 's8'], ['s7', 's1'], ['s7'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3', 's1'], ['s7', 's2'], ['s3', 's7'], ['s1', 's2'], ['s1', 's2'], ['s1', 's7'],
                  ['s2', 's3'], ['s3', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s2'], ['s1', 's2'], ['s4'], ['s2', 's4'], ['s1', 's4'],
                  ['s2', 's1']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s5'], ['s5', 's1'], ['s2'], ['s2'], ['s5'], ['s2', 's5'], ['s1'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5', 's3'], ['s3', 's2'], ['s1', 's5'], ['s5', 's2'], ['s1', 's2'], ['s5', 's1'], ['s3', 's2'], ['s3'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7'], ['s7'], ['s1'], ['s1', 's2'], ['s1'], ['s2'], ['s1'], ['s7', 's2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's8'], ['s5', 's1', 's7'], ['s4', 's5'], ['s4', 's7'], ['s5', 's8'], ['s5', 's7'], ['s1', 's8', 's4'], ['s1', 's7'],
                  ['s4', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5'], ['s10'], ['s5', 's4'], ['s1'], ['s10', 's5'], ['s1', 's4'], ['s10', 's4'], ['s5', 's1'], ['s4'],
                  ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1'], ['s7', 's9'], ['s8', 's1'], ['s9', 's5'], ['s1', 's5'], ['s8', 's5'], ['s7', 's5'],
                  ['s9', 's1'], ['s8', 's9'], ['s7', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1'], ['s7', 's9', 's4'], ['s8', 's1'], ['s9', 's2', 's1'], ['s2', 's4'], ['s8', 's4'],
                  ['s4', 's8', 's2'],
                  ['s7', 's9'], ['s2', 's9'], ['s1', 's8', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1', 's6'], ['s7', 's9', 's4'], ['s8', 's1'], ['s9', 's2', 's6'], ['s2', 's6', 's1'], ['s8', 's4'],
                  ['s4', 's8', 's6'],
                  ['s7', 's9', 's4'], ['s2', 's9', 's7'], ['s1', 's8', 's2']])])

    elif vehicle_number == -10:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s1'], ['s1', 's2'], ['s2', 's1'], ['s2', 's1'], ['s2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1'], ['s1', 's6'], ['s6', 's1'], ['s1', 's6'], ['s6', 's1'],
                  ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's7'], ['s7', 's8', 's1'], ['s8', 's7', 's1'], ['s7', 's1', 's8'], ['s1', 's8'], ['s1', 's7']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5'],
                 #[['s1', 's8'], ['s1', 's8'], ['s1', 's8'], ['s8', 's1'], ['s8', 's1']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5'],
                 #[['s1'], ['s1'], ['s1'], ['s1'], ['s1']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5'],
                 #[['s1', 's2', 's8'], ['s2', 's8', 's1'], ['s8', 's1', 's2'], ['s8', 's2', 's1'], ['s1', 's2', 's8']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5'],
                #  [['s2', 's7', 's1'], ['s1', 's2', 's5'], ['s7', 's1', 's2'], ['s5', 's7', 's2'], ['s1', 's5', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5', 's8'], ['s8', 's4', 's1'], ['s1', 's8', 's4'], ['s4', 's1', 's5'], ['s1', 's5', 's8'], ['s5', 's8'], ['s4', 's1', 's5']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s9'], ['s9'], ['s9'], ['s9', 's4'], ['s4', 's9'], ['s4'], ['s4'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's7'], ['s1', 's7'], ['s8', 's1'], ['s7', 's1'], ['s1', 's8'], ['s7', 's1'], ['s7', 's8'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3', 's1'], ['s7', 's2'], ['s3', 's7', 's1'], ['s1', 's2', 's3'], ['s1', 's2', 's7'], ['s1', 's7'],
                  ['s2', 's3'], ['s3', 's7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s2', 's4'], ['s1', 's2'], ['s4', 's1', 's2'], ['s2', 's4'], ['s1', 's4'],
                  ['s2', 's1']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s5', 's1'], ['s5', 's1'], ['s2', 's5'], ['s2'], ['s5'], ['s2', 's5'], ['s1', 's2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1'], ['s1'], ['s2'], ['s2'], ['s2'], ['s2'], ['s1'],
                  ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5', 's3'], ['s3', 's2'], ['s1', 's5'], ['s5', 's2'], ['s1', 's2', 's5'], ['s5', 's1', 's3'], ['s3', 's2'], ['s3', 's1'], ['s1', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's2'], ['s7', 's1'], ['s1'], ['s1', 's2'], ['s1', 's7'], ['s2'], ['s1'], ['s7', 's2'],
                  ['s2']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 #[['s7'], ['s7'], ['s7'], ['s1'], ['s1'], ['s1'], ['s7'], ['s1'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's8', 's7'], ['s5', 's1', 's7'], ['s4', 's5'], ['s4', 's7', 's8'], ['s5', 's8', 's4'], ['s5', 's7'], ['s1', 's8', 's4'], ['s1', 's7', 's5'],
                  ['s4', 's8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5', 's4'], ['s10', 's5'], ['s5', 's4'], ['s1', 's10'], ['s10', 's5'], ['s1', 's4'], ['s10', 's4'], ['s5', 's1'], ['s4', 's1'],
                  ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1', 's9'], ['s7', 's9'], ['s8', 's1', 's5'], ['s9', 's5', 's8'], ['s1', 's5'], ['s8', 's5'], ['s7', 's5'],
                  ['s9', 's1'], ['s8', 's9', 's7'], ['s7', 's8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1', 's9'], ['s7', 's9', 's4'], ['s8', 's1', 's4'], ['s9', 's2', 's1'], ['s2', 's4', 's8'], ['s8', 's4'],
                  ['s4', 's8', 's2'],
                  ['s7', 's9', 's2'], ['s2', 's9', 's1'], ['s1', 's8', 's7']])])
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                #  [['s7', 's1', 's6'], ['s7', 's9', 's4'], ['s8', 's1'], ['s9', 's2', 's6'], ['s2', 's6', 's1'], ['s8', 's4'],
                #   ['s4', 's8', 's6'],
                #   ['s7', 's9', 's4'], ['s2', 's9', 's7'], ['s1', 's8', 's2']])])

    elif vehicle_number == -11:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1'], ['s1', 's2'], ['s1', 's2'], ['s2', 's1'], ['s2', 's1'], ['s1', 's2'], ['s2']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 #[['s1', 's6'], ['s1', 's6'], ['s6', 's1'], ['s1', 's6'], ['s6', 's1'],
                  #['s6', 's1']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 #[['s1'], ['s1'], ['s1'], ['s1'], ['s1'],
                  #['s1']]),
                #(['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 #[['s8', 's7', 's1'], ['s7', 's8', 's1'], ['s8', 's7', 's1'], ['s7', 's1', 's8'], ['s1', 's8', 's7'], ['s1', 's7', 's8']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5'],
                #  [['s1', 's8'], ['s1', 's8'], ['s1', 's8'], ['s8', 's1'], ['s8', 's1']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5'],
                #  [['s1'], ['s1'], ['s1'], ['s1'], ['s1']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5'],
                #  [['s1', 's2', 's8'], ['s2', 's8', 's1'], ['s8', 's1', 's2'], ['s8', 's2', 's1'], ['s1', 's2', 's8']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5'],
                #  [['s2', 's7', 's1'], ['s1', 's2', 's5'], ['s7', 's1', 's2'], ['s5', 's7', 's2'], ['s1', 's5', 's7']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                #  [['s4', 's5', 's8'], ['s8', 's4', 's1'], ['s1', 's8', 's4'], ['s4', 's1', 's5'], ['s1', 's5', 's8'], ['s5', 's8'], ['s4', 's1', 's5']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s9'], ['s9'], ['s9', 's4'], ['s9', 's4'], ['s4', 's9'], ['s4', 's9'], ['s4'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's7', 's1'], ['s1', 's7'], ['s8', 's1'], ['s7', 's1'], ['s1', 's8'], ['s7', 's1', 's8'], ['s7', 's8'], ['s8', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3', 's1', 's2'], ['s7', 's2', 's1'], ['s3', 's7', 's1'], ['s1', 's2', 's3'], ['s1', 's2', 's7'], ['s1', 's7', 's3'],
                  ['s2', 's3', 's7'], ['s3', 's7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4', 's2'], ['s2', 's4', 's1'], ['s1', 's2'], ['s4', 's1', 's2'], ['s2', 's4'], ['s1', 's4'],
                  ['s2', 's1', 's4']]),

                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s5', 's1'], ['s5', 's1'], ['s2', 's5'], ['s2', 's1'], ['s5', 's2'], ['s2', 's5'], ['s1', 's2'], ['s1', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s1', 's2'], ['s1'], ['s2', 's1'], ['s2'], ['s2'], ['s2'], ['s1'],
                  ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5', 's3'], ['s3', 's2', 's5'], ['s1', 's5', 's3'], ['s5', 's2', 's1'], ['s1', 's2', 's5'], ['s5', 's1', 's3'], ['s3', 's2'], ['s3', 's1', 's2'], ['s1', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's2'], ['s7', 's2'], ['s7', 's1'], ['s1'], ['s1', 's2'], ['s1', 's7'], ['s2', 's1'], ['s1', 's7'], ['s7', 's2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7'], ['s7', 's1'], ['s1'], ['s1', 's7'], ['s1'], ['s7'], ['s1'], ['s7'], ['s1']]),
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                #  [['s1', 's8', 's7'], ['s5', 's1', 's7'], ['s4', 's5'], ['s4', 's7', 's8'], ['s5', 's8', 's4'], ['s5', 's7'], ['s1', 's8', 's4'], ['s1', 's7', 's5'],
                #   ['s4', 's8', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5', 's4', 's10'], ['s10', 's5', 's1'], ['s5', 's4'], ['s1', 's10', 's4'], ['s10', 's5'], ['s1', 's4', 's5'], ['s10', 's4'], ['s5', 's1'], ['s4', 's1'],
                  ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's1', 's9'], ['s7', 's9', 's5'], ['s8', 's1', 's5'], ['s9', 's5', 's8'], ['s1', 's5', 's8'], ['s8', 's5', 's7'], ['s7', 's5', 's9'],
                  ['s9', 's1', 's8'], ['s8', 's9', 's7'], ['s7', 's8', 's1']])])
                # (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                #  [['s7', 's1', 's9'], ['s7', 's9', 's4'], ['s8', 's1', 's4'], ['s9', 's2', 's1'], ['s2', 's4', 's8'], ['s8', 's4'],
                #   ['s4', 's8', 's2'],
                #   ['s7', 's9', 's2'], ['s2', 's9', 's1'], ['s1', 's8', 's7']])])


    elif vehicle_number == 4:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's7'], ['s3', 's1', 's6'], ['s3', 's2', 's7'], ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's6'], ['s8', 's5', 's6'], ['s5', 's1', 's2'], ['s5', 's1']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s8', 's4'], ['s7', 's8'], ['s7', 's4', 's2'], ['s8', 's4', 's2']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s4', 's1', 's7'], ['s2', 's4'], ['s2', 's1', 's4'], ['s7', 's1']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's3'], ['s3', 's2', 's1'], ['s7', 's2', 's8'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s8', 's1', 's2'], ['s1', 's4'], ['s1', 's8'], ['s2', 's8', 's4']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's3', 's8'], ['s2', 's6', 's5'], ['s2', 's5'], ['s8', 's1']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s2', 's6', 's7'], ['s1', 's6', 's7'], ['s7', 's5'], ['s8', 's2']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s4', 's5', 's8'], ['s1', 's5', 's4'], ['s8', 's1'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s7', 's4', 's6'], ['s1', 's4'], ['s6', 's4'], ['s7', 's6', 's1']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4'],
                 [['s9', 's7', 's1'], ['s4', 's8'], ['s9', 's1'], ['s8', 's4', 's2']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s3', 's9'], ['s9', 's3'], ['s6', 's8', 's2'], ['s8', 's2']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s3', 's4', 's2'], ['s1', 's6', 's7'], ['s4', 's2'], ['s5', 's4']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's4'], ['s3', 's2', 's5'], ['s4', 's6', 's1'], ['s5', 's2']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s4', 's7', 's1'], ['s7', 's4', 's6'], ['s2', 's8'], ['s8', 's9']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s4', 's8'], ['s7', 's1'], ['s4', 's8', 's2'], ['s7', 's3', 's1']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8', 's1'], ['s7', 's8', 's1']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's8'], ['s4', 's8', 's1'], ['s5', 's7', 's2'], ['s6', 's5']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1', 's2'], ['s5', 's7']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s3', 's6', 's4'], ['s4', 's6'], ['s1', 's2', 's5'], ['s8', 's1']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s5', 's9']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's5'], ['s4', 's9'], ['s2', 's3'], ['s3', 's7']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's3'], ['s10', 's1'], ['s5'], ['s8', 's7']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s7'], ['s4'], ['s3'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9']])])
        # con 4 subflotas
        # elif subfleet_number == 4:
        #     return random.choice([
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s1'], ['s9'], ['s1'], ['s6'], ['s7'], ['s9'], ['s9'], ['s9']]),
        #         (['v1', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s4'], ['s1'], ['s3'], ['s2'], ['s1', 's5'], ['s10'], ['s10', 's2']]),
        #         (['v1', 'v2', 'v4', 'v5', 'v7', 'v8', 'v9'],
        #          [['s1'], ['s2'], ['s4', 's6', 's8'], ['s8'], ['s7'], ['s8'], ['s2']]),
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s4'], ['s6'], ['s3'], ['s4'], ['s6'], ['s6'], ['s8'], ['s3'], ['s8']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v10'],
        #          [['s1'], ['s8', 's6'], ['s10'], ['s10', 's10'], ['s10'], ['s6'], ['s7'], ['s10'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s3'], ['s3', 's1'], ['s4'], ['s9'], ['s3'], ['s9'], ['s9'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s6'], ['s3'], ['s3'], ['s6'], ['s6'], ['s10'], ['s3'], ['s3'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v7', 'v8', 'v9', 'v10'],
        #          [['s6', 's3'], ['s9'], ['s3'], ['s6'], ['s8'], ['s7'], ['s8'], ['s9'], ['s3']]),
        #         (['v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s2'], ['s2'], ['s9'], ['s10'], ['s6'], ['s7'], ['s7'], ['s9'], ['s6', 's7']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v7', 'v8', 'v9', 'v10'],
        #          [['s6', 's7'], ['s2'], ['s9'], ['s7'], ['s6'], ['s7'], ['s7'], ['s9'], ['s10']])])

    elif vehicle_number == 5:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s8', 's4'], ['s7', 's8'], ['s7', 's4'], ['s8', 's2'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's1'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's3'], ['s3', 's2'], ['s2', 's8'], ['s8', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s2', 's4'], ['s1', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's3'], ['s6', 's5'], ['s2', 's5'], ['s8', 's1'], ['s2', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s2', 's6'], ['s1', 's7'], ['s7', 's5'], ['s8', 's2'], ['s6', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7', 's1'], ['s1', 's6']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s9', 's7', 's1'], ['s4', 's8'], ['s9', 's1'], ['s8', 's4'], ['s2', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s6'], ['s9', 's3'], ['s6', 's8', 's2'], ['s8', 's2'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s3', 's4', 's2'], ['s1', 's6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's4'], ['s3', 's2', 's5'], ['s4', 's6'], ['s5', 's2'], ['s6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s4', 's7', 's1'], ['s7', 's4', 's6'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s4', 's8'], ['s7', 's1'], ['s4', 's8', 's2'], ['s7', 's3', 's1'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8', 's1'], ['s7', 's8', 's1'], ['s4', 's5', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's8'], ['s4', 's8', 's1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1', 's2'], ['s5', 's7'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s3', 's6', 's4'], ['s4', 's6'], ['s1', 's2', 's5'], ['s8', 's1'], ['s3', 's6']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s5', 's9'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s9'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s6', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's5'], ['s4', 's9'], ['s2', 's3'], ['s3', 's7'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's3'], ['s10', 's1'], ['s5'], ['s8', 's7'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4', 's7'], ['s4', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9']])])

    elif vehicle_number == 6:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6', 's7'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s2', 's1'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8', 's2'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's1'], ['s4', 's7'], ['s4', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's3'], ['s3', 's2'], ['s2', 's8'], ['s8', 's1'], ['s1', 's7'], ['s7', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s2'], ['s4'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's3'], ['s6', 's5'], ['s2'], ['s8', 's1'], ['s2', 's8'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s2', 's6'], ['s1', 's7'], ['s7', 's5'], ['s8', 's2'], ['s6'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8', 's5'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7', 's1'], ['s1', 's6'], ['s6', 's9']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8', 's4'], ['s2', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s6'], ['s9', 's3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s3', 's4', 's2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's4'], ['s3', 's2', 's5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s4'], ['s7', 's1'], ['s7', 's4', 's6'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4', 's8', 's2'], ['s7', 's3', 's1'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7', 's8', 's1'], ['s4', 's5', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1', 's2'], ['s5', 's7'], ['s4'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2', 's5'], ['s8', 's1'], ['s3', 's6']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s5'], ['s5'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s8'], ['s9'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s6'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8']])])

    elif vehicle_number == 7:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s8', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's3'], ['s3', 's2'], ['s2', 's8'], ['s8'], ['s1'], ['s1', 's7'], ['s7', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'],  ['s2'], ['s4'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's3'], ['s6', 's5'], ['s2'], ['s8'], ['s1'], ['s2', 's8'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s2', 's6'], ['s1', 's7'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8'], ['s4'], ['s8', 's5'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s1'], ['s1', 's6'], ['s6', 's9']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s4'], ['s2', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s3', 's6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4', 's5', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2']])])

    elif vehicle_number == 8:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'], ['s1', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'], ['s8', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s2', 's4'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's3'], ['s3', 's2'], ['s2', 's8'], ['s8'], ['s1'], ['s1', 's7'], ['s7', 's4'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s4'], ['s8'], ['s1', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's3'], ['s6', 's5'], ['s2'], ['s8'], ['s1'], ['s2', 's8'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s2', 's6'], ['s1', 's7'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s7'], ['s6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8'], ['s4'], ['s8', 's5'], ['s1'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s9']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s4'], ['s2', 's1'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s3', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s5', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s3', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s1'], ['s5', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1']])])

    elif vehicle_number == 9:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'],
                  ['s1', 's3'], ['s3', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'],
                  ['s8', 's6'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s2', 's4'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2'],
                  ['s2'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's3'], ['s3', 's2'], ['s2', 's8'], ['s8'], ['s1'], ['s1', 's7'], ['s7', 's4'],
                  ['s4'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s4'], ['s8'],
                  ['s1', 's2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s8'], ['s1'], ['s2', 's8'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s7'], ['s6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s4'], ['s8', 's5'], ['s1'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'],
                  ['s9']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s2', 's1'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s7'], ['s9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'],
                  ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'],
                  ['s3', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'],
                  ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'],
                  ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'],
                  ['s5', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'],
                  ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s4', 's7'],
                  ['s3', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2'],
                  ['s2']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s5', 's10'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5'], ['s2', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10'],
                  ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5']])])















        # if subfleet_number == 1:
        #     return random.choice([
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v7', 'v8'],
        #          [['s1', 's6'], ['s3', 's1'], ['s3'], ['s4'], ['s1'], ['s4', 's3'], ['s6']]),
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7'],
        #          [['s1', 's6'], ['s8', 's5'], ['s5'], ['s5', 's1'], ['s6'],
        #           ['s8', 's6']]),
        #         (['v3', 'v6', 'v7', 'v8'],
        #          [['s8', 's4'], ['s7', 's8'], ['s7', 's4', 's2'], ['s8', 's4', 's2']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v6', 'v7'],
        #          [['s1'], ['s4', 's1'], ['s2', 's4'], ['s4'], ['s2', 's1'], ['s7', 's1']]),
        #         (['v1', 'v3', 'v5', 'v6', 'v7'],
        #          [['s1'], ['s3'], ['s3', 's2', 's1'], ['s7', 's2', 's8'], ['s8', 's1']]),
        #         (['v2', 'v3', 'v4', 'v5', 'v7', 'v8'],
        #          [['s2'], ['s8', 's1'], ['s4'], ['s1', 's4'], ['s1', 's8'], ['s2', 's8']]),
        #         (['v1', 'v2', 'v6', 'v7', 'v8'],
        #          [['s1', 's3'], ['s5', 's8'], ['s2', 's6'], ['s2', 's5'], ['s8', 's1']]),
        #         (['v2', 'v4', 'v5', 'v7', 'v8'],
        #          [['s2', 's6', 's7'], ['s1', 's6'], ['s7'], ['s7', 's5'], ['s8', 's2']]),
        #         (['v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
        #          [['s4', 's5'], ['s4'], ['s5'], ['s8'], ['s1'], ['s8', 's1'], ['s8', 's4']]),
        #         (['v1', 'v2', 'v3', 'v5', 'v7', 'v8'],
        #          [['s7', 's4'], ['s1'], ['s4'], ['s6', 's4'], ['s7', 's6'], ['s1', 's6']])])
        # # con 2 subflotas
        # elif subfleet_number == 2:
        #     return random.choice([
        #         (['v1', 'v2', 'v3', 'v5', 'v9', 'v10', 'v12', 'v13'],
        #          [['s9'], ['s4'], ['s8'], ['s9', 's1'], ['s1'], ['s4'], ['s9'], ['s8', 's4']]),
        #         (['v1', 'v3', 'v4', 'v6', 'v8', 'v10'],
        #          [['s9'], ['s3', 's6', 's9'], ['s9', 's3'], ['s6'], ['s8'], ['s8', 's2']]),
        #         (['v2', 'v3', 'v4', 'v5'],
        #          [['s3', 's4', 's2'], ['s1', 's6', 's7'], ['s4', 's2'], ['s5', 's4']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5'],
        #          [['s1', 's4'], ['s2'], ['s3', 's2', 's5'], ['s4', 's6'], ['s5', 's7']]),
        #         (['v1', 'v2', 'v5', 'v7', 'v8'],
        #          [['s1'], ['s4', 's7', 's1'], ['s7', 's4'], ['s2', 's8'], ['s8', 's9']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
        #          [['s8'], ['s7'], ['s4'], ['s4', 's8'], ['s4'], ['s4', 's8'], ['s7'], ['s8']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8'],
        #          [['s3'], ['s8'], ['s3'], ['s4', 's7'], ['s4', 's3'], ['s7'], ['s7'], ['s8']]),
        #         (['v1', 'v2', 'v5', 'v6', 'v7', 'v8'],
        #          [['s1', 's8'], ['s4', 's8'], ['s1', 's4'], ['s6'], ['s6', 's5'], ['s4']]),
        #         (['v1', 'v2', 'v4', 'v5', 'v6', 'v7', 'v8'],
        #          [['s1'], ['s2', 's6'], ['s1'], ['s3', 's4'], ['s2'], ['s3', 's1'], ['s2']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v6', 'v7', 'v8'],
        #          [['s1'], ['s3'], ['s3', 's6', 's4'], ['s4'], ['s6'], ['s1'], ['s8', 's1']])])
        # # con 3 subflotas
        # elif subfleet_number == 3:
        #     return random.choice([
        #         (['v3', 'v4', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s3', 's1'], ['s6', 's10'], ['s3'], ['s7'], ['s5', 's3'], ['s3'], ['s3']]),
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v9'],
        #          [['s1'], ['s1', 's5'], ['s10'], ['s5'], ['s6', 's8'], ['s5', 's10'], ['s9']]),
        #         (['v1', 'v3', 'v4', 'v5', 'v6', 'v7', 'v9', 'v10'],
        #          [['s2'], ['s10'], ['s7', 's10'], ['s5'], ['s10'], ['s10', 's8'], ['s7'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s5'], ['s5'], ['s4'], ['s9'], ['s9'], ['s5'], ['s9', 's5'], ['s9']]),
        #         (['v1', 'v2', 'v4', 'v5', 'v6', 'v8', 'v9', 'v10'],
        #          [['s1', 's3'], ['s10'], ['s1'], ['s5'], ['s5'], ['s8', 's1'], ['s10'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v9', 'v10'],
        #          [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s7'], ['s9', 's3'], ['s7', 's1']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v9', 'v10'],
        #          [['s5'], ['s10'], ['s5', 's8'], ['s2'], ['s10'], ['s10'], ['s8'], ['s8'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1', 's2'], ['s2', 's10'], ['s6'], ['s4'], ['s1'], ['s6'], ['s4'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v9', 'v10'],
        #          [['s1'], ['s2'], ['s3'], ['s1', 's6'], ['s4'], ['s4'], ['s4', 's3'], ['s2']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s7'], ['s7'], ['s8'], ['s9'], ['s9'], ['s8'], ['s8'], ['s8'], ['s8'], ['s7']])])
        # # con 4 subflotas
        # elif subfleet_number == 4:
        #     return random.choice([
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s1'], ['s9'], ['s1'], ['s6'], ['s7'], ['s9'], ['s9'], ['s9']]),
        #         (['v1', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s4'], ['s1'], ['s3'], ['s2'], ['s1', 's5'], ['s10'], ['s10', 's2']]),
        #         (['v1', 'v2', 'v4', 'v5', 'v7', 'v8', 'v9'],
        #          [['s1'], ['s2'], ['s4', 's6', 's8'], ['s8'], ['s7'], ['s8'], ['s2']]),
        #         (['v1', 'v2', 'v3', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s4'], ['s6'], ['s3'], ['s4'], ['s6'], ['s6'], ['s8'], ['s3'], ['s8']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v10'],
        #          [['s1'], ['s8', 's6'], ['s10'], ['s10', 's10'], ['s10'], ['s6'], ['s7'], ['s10'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s3'], ['s3', 's1'], ['s4'], ['s9'], ['s3'], ['s9'], ['s9'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s1'], ['s6'], ['s3'], ['s3'], ['s6'], ['s6'], ['s10'], ['s3'], ['s3'], ['s10']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v7', 'v8', 'v9', 'v10'],
        #          [['s6', 's3'], ['s9'], ['s3'], ['s6'], ['s8'], ['s7'], ['s8'], ['s9'], ['s3']]),
        #         (['v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
        #          [['s2'], ['s2'], ['s9'], ['s10'], ['s6'], ['s7'], ['s7'], ['s9'], ['s6', 's7']]),
        #         (['v1', 'v2', 'v3', 'v4', 'v5', 'v7', 'v8', 'v9', 'v10'],
        #          [['s6', 's7'], ['s2'], ['s9'], ['s7'], ['s6'], ['s7'], ['s7'], ['s9'], ['s10']])])

    elif vehicle_number == 10:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'],
                  ['s8', 's6'], ['s8'], ['s8', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s2', 's4'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2'],
                  ['s2'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's3'], ['s3', 's2'], ['s2', 's8'], ['s8'], ['s1', 's8'], ['s1'], ['s1', 's7'], ['s7', 's4'],
                  ['s4'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s4', 's2'], ['s4'], ['s8'],
                  ['s1', 's2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s8'], ['s1'], ['s2', 's8'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s7'], ['s6', 's7'], ['s6', 's1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s4'], ['s8', 's5'], ['s1'], ['s5'], ['s5', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s7', 's4'],
                  ['s9']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s2', 's1'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s7'], ['s9'], ['s8', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'],
                  ['s5'], ['s7', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'],
                  ['s3', 's5'], ['s2', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'],
                  ['s2'], ['s8', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'],
                  ['s2', 's4'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'],
                  ['s5', 's3'], ['s5', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'],
                  ['s4'], ['s1', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s4', 's7'],
                  ['s3', 's6'], ['s1', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2'],
                  ['s2'], ['s4', 's3']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6'], ['s7'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2'], ['s10'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4'], ['s10'], ['s7', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s5', 's10'], ['s10'], ['s8', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5'], ['s2', 's10'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s3', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10'],
                  ['s2', 's6'], ['s1', 's2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'], ['s2', 's9']])])

    elif vehicle_number == 11:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s1'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s2', 's4'], ['s7', 's2'], ['s1', 's2', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2'],
                  ['s2'], ['s4'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's3'], ['s3', 's2'], ['s2', 's8'], ['s8'], ['s1', 's8'], ['s1'], ['s1', 's7'], ['s7', 's4'],
                  ['s4'], ['s8', 's4'], ['s3', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s4', 's2'], ['s4'], ['s8'],
                  ['s1', 's2'], ['s1', 's3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s5'], ['s8'], ['s1'], ['s2', 's8'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s7'], ['s6', 's7'], ['s6', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s4'], ['s8', 's5'], ['s1'], ['s5'], ['s5', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s7', 's4'],
                  ['s9'], ['s1', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s2', 's1'], ['s2'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s7'], ['s9'], ['s8', 's2'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'],
                  ['s5'], ['s7', 's6'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'],
                  ['s3', 's5'], ['s2', 's3'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'],
                  ['s2'], ['s8', 's9'], ['s1', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'],
                  ['s2', 's4'], ['s2'], ['s4', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'],
                  ['s5', 's3'], ['s5', 's3'], ['s3', 's4', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'],
                  ['s4'], ['s1', 's8'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s4', 's7'],
                  ['s3', 's6'], ['s1', 's2'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2'],
                  ['s2'], ['s4', 's3'], ['s3']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6'], ['s7'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2'], ['s10'], ['s6'], ['s5', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4'], ['s10'], ['s7', 's10'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s5', 's10'], ['s10'], ['s8', 's4'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3'], ['s1', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5'], ['s2', 's10'], ['s5'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s3', 's5'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10'],
                  ['s2', 's6'], ['s1', 's2', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'], ['s2', 's9'], ['s8']])])

    elif vehicle_number == 12:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s1'], ['s6'], ['s8', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'],
                  ['s2', 's4'], ['s7', 's2'], ['s1', 's2', 's7'], ['s1', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2'],
                  ['s2'], ['s4'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's3'], ['s3', 's2'], ['s1', 's2'], ['s2', 's8'], ['s8'], ['s1', 's8'], ['s1'], ['s1', 's7'], ['s7', 's4'],
                  ['s4'], ['s8', 's4'], ['s3', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s4', 's2'], ['s4'], ['s8'],
                  ['s1', 's2'], ['s1', 's3'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s5'], ['s8'], ['s1'], ['s2', 's8'], ['s1', 's6'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s7'], ['s2', 's5'], ['s6', 's7'], ['s6', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s1', 's8'], ['s4'], ['s8', 's5'], ['s1'], ['s5'], ['s5', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s7', 's4'], ['s2'],
                  ['s9'], ['s1', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s2', 's1'], ['s4', 's8'], ['s2'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s7'], ['s9'], ['s8', 's2'], ['s7'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'], ['s3'],
                  ['s5'], ['s7', 's6'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'], ['s2'],
                  ['s3', 's5'], ['s2', 's3'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'], ['s7'],
                  ['s2'], ['s8', 's9'], ['s1', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'], ['s8'],
                  ['s2', 's4'], ['s2'], ['s4', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'], ['s3', 's4'],
                  ['s5', 's3'], ['s5', 's3'], ['s3', 's4', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'], ['s2', 's6'],
                  ['s4'], ['s1', 's8'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s4', 's7'], ['s6'],
                  ['s3', 's6'], ['s1', 's2'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2'], ['s3', 's6'],
                  ['s2'], ['s4', 's3'], ['s3']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6'], ['s7'], ['s3'], ['s3'], ['s6', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2'], ['s10'], ['s6'], ['s5', 's10'], ['s2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4'], ['s10'], ['s7', 's10'], ['s5'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3'], ['s3'], ['s4', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s5', 's10'], ['s10'], ['s8', 's4'], ['s8'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3'], ['s1', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5'], ['s2', 's10'], ['s5'], ['s8'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s3', 's5'], ['s2'], ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10'], ['s10'],
                  ['s2', 's6'], ['s1', 's2', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'], ['s2', 's9'], ['s8'], ['s7']])])

    elif vehicle_number == 13:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6'], ['s1', 's7'], ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s6'], ['s8', 's5'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'],
                  ['s2', 's4'], ['s7', 's2'], ['s1', 's2', 's7'], ['s1', 's4'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2'],
                  ['s2'], ['s4'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's3'], ['s3', 's2'], ['s1', 's2'], ['s2', 's8'], ['s8'], ['s7'], ['s1', 's8'], ['s1'], ['s1', 's7'], ['s7', 's4'],
                  ['s4'], ['s8', 's4'], ['s3', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s2', 's3'], ['s4', 's2'], ['s4'], ['s8'],
                  ['s1', 's2'], ['s1', 's3'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s5'], ['s8'], ['s1'], ['s3'], ['s2', 's8'], ['s1', 's6'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s5'], ['s7'], ['s2', 's5'], ['s6', 's7'], ['s6', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s1', 's8'], ['s4'], ['s2', 's5'],['s8', 's5'], ['s1'], ['s5'], ['s5', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s1'], ['s7', 's4'], ['s2'],
                  ['s9'], ['s2', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s9'], ['s2', 's1'], ['s4', 's8'], ['s2'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s6'], ['s7'], ['s9'], ['s8', 's2'], ['s7'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'], ['s3'], ['s1', 's7'],
                  ['s5'], ['s7', 's6'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'], ['s2'], ['s6'],
                  ['s3', 's5'], ['s2', 's3'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'], ['s7'], ['s1', 's4'],
                  ['s2'], ['s8', 's9'], ['s1', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'], ['s8'], ['s1', 's7'],
                  ['s2', 's4'], ['s2'], ['s4', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'], ['s3', 's4'], ['s1', 's8'],
                  ['s5', 's3'], ['s5', 's3'], ['s3', 's4', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'], ['s2', 's6'], ['s1', 's4'],
                  ['s4'], ['s1', 's8'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s4', 's7'], ['s6'], ['s5'],
                  ['s3', 's6'], ['s1', 's2'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2'], ['s3', 's6'], ['s1'],
                  ['s2'], ['s4', 's3'], ['s3']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6'], ['s7'], ['s3'], ['s3'], ['s6', 's10'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2'], ['s10'], ['s6'], ['s5', 's10'], ['s2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4'], ['s10'], ['s7', 's10'], ['s5'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3'], ['s3'], ['s4', 's9'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s5', 's10'], ['s10'], ['s8', 's4'], ['s8'], ['s4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3'], ['s1', 's4'], ['s5'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5'], ['s2', 's10'], ['s5'], ['s8'], ['s7'], ['s10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s3', 's5'], ['s2'], ['s10'], ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10'], ['s10'], ['s1', 's2'],
                  ['s2', 's6'], ['s1', 's2', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'], ['s2', 's9'], ['s8'], ['s5'], ['s7']])])

    elif vehicle_number == 14:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6'], ['s1', 's7'], ['s2', 's6'], ['s1', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'], ['s1', 's6'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s6'], ['s8', 's5'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s1'],
                  ['s2', 's4'], ['s7', 's2'], ['s1', 's2', 's7'], ['s1', 's4'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s7'], ['s1'], ['s4', 's7'], ['s4', 's2'], ['s2', 's7'],
                  ['s2'], ['s4'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's3'], ['s3', 's2'], ['s1', 's2'], ['s2', 's8'], ['s8'], ['s7'], ['s1', 's8'], ['s1'], ['s1', 's7'], ['s7', 's4'], ['s7'],
                  ['s4'], ['s8', 's4'], ['s3', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s2', 's3'], ['s4', 's2'], ['s4'], ['s8'], ['s3'],
                  ['s1', 's2'], ['s1', 's3'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s5'], ['s8'], ['s1'], ['s3'], ['s2', 's8'], ['s1', 's6'], ['s8'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s5'], ['s7'], ['s2', 's5'], ['s5'], ['s6', 's7'], ['s6', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s1', 's8'], ['s4'], ['s2', 's5'],['s8', 's5'], ['s8'], ['s1'], ['s5'], ['s5', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s1'], ['s7', 's4'], ['s2'], ['s2'],
                  ['s9'], ['s2', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s9'], ['s2', 's1'], ['s4'], ['s4', 's8'], ['s2'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s6'], ['s7'], ['s9'], ['s7'], ['s8', 's2'], ['s7'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'], ['s3'], ['s1', 's7'], ['s5'],
                  ['s5'], ['s7', 's6'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'], ['s2'], ['s6'], ['s5'],
                  ['s3', 's5'], ['s2', 's3'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'], ['s7'], ['s1', 's4'], ['s8'],
                  ['s2'], ['s8', 's9'], ['s1', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'], ['s8'], ['s1', 's7'], ['s2'],
                  ['s2', 's4'], ['s2'], ['s4', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'], ['s3', 's4'], ['s1', 's8'], ['s7'],
                  ['s5', 's3'], ['s5', 's3'], ['s3', 's4', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'], ['s2', 's6'], ['s1', 's4'], ['s7'],
                  ['s4'], ['s1', 's8'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s4', 's7'], ['s6'], ['s5'], ['s3'],
                  ['s3', 's6'], ['s1', 's2'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2'], ['s3', 's6'], ['s1'], ['s8'],
                  ['s2'], ['s4', 's3'], ['s3']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6'], ['s7'], ['s3'], ['s5'], ['s3'], ['s6', 's10'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2'], ['s10'], ['s6'], ['s2'], ['s5', 's10'], ['s2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4'], ['s10'], ['s7', 's10'], ['s5'], ['s8'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3'], ['s3'], ['s4', 's9'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s5', 's10'], ['s10'], ['s8', 's4'], ['s8'], ['s4'], ['s4', 's7'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3'], ['s1', 's4'], ['s5'], ['s3'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5'], ['s2', 's10'], ['s5'], ['s8'], ['s7'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s3', 's5'], ['s2'], ['s10'], ['s3'], ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10'], ['s10'], ['s1', 's2'], ['s5'],
                  ['s2', 's6'], ['s1', 's2', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'], ['s2', 's9'], ['s8'], ['s5'], ['s2'], ['s7']])])

    elif vehicle_number == 15:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6'], ['s1', 's7'], ['s2', 's6'], ['s1', 's3'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'], ['s1', 's6'], ['s5'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s6'], ['s8', 's5'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s1'], ['s8'],
                  ['s2', 's4'], ['s7', 's2'], ['s1', 's2', 's7'], ['s1', 's4'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s7'], ['s1'], ['s4', 's7'], ['s1'], ['s4', 's2'], ['s2', 's7'],
                  ['s2'], ['s4'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's3'], ['s3', 's2'], ['s1', 's2'], ['s2', 's8'], ['s8'], ['s7'], ['s1', 's8'], ['s1'], ['s2', 's3'], ['s1', 's7'], ['s7', 's4'], ['s7'],
                  ['s4'], ['s8', 's4'], ['s3', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s2', 's3'], ['s4', 's2'], ['s4'], ['s8'], ['s1', 's4'], ['s3'],
                  ['s1', 's2'], ['s1', 's3'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s5'], ['s8'], ['s1'], ['s3'], ['s2', 's8'], ['s1', 's6'], ['s8'], ['s2', 's6'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s5'], ['s7'], ['s2', 's5'], ['s5'], ['s6', 's7'], ['s2'], ['s6', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s1', 's8'], ['s4'], ['s2', 's5'],['s8', 's5'], ['s8'], ['s1'], ['s5'], ['s1'], ['s5', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s1'], ['s7', 's4'], ['s2'], ['s2'], ['s6', 's9'],
                  ['s9'], ['s2', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s9'], ['s7', 's9'], ['s2', 's1'], ['s4'], ['s4', 's8'], ['s2'], ['s7', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s6'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s6'], ['s7'], ['s6', 's8'], ['s9'], ['s7'], ['s8', 's2'], ['s7'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s3', 's4'], ['s2'], ['s1'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'], ['s3'], ['s3', 's4'], ['s1', 's7'], ['s5'],
                  ['s5'], ['s7', 's6'], ['s2', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'], ['s2'], ['s6'], ['s5'], ['s1', 's4'],
                  ['s3', 's5'], ['s2', 's3'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'], ['s7'], ['s1', 's4'], ['s8'], ['s2', 's9'],
                  ['s2'], ['s8', 's9'], ['s1', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'], ['s8'], ['s1', 's7'], ['s2'], ['s7'],
                  ['s2', 's4'], ['s2'], ['s4', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'], ['s3', 's4'], ['s1', 's8'], ['s7'], ['s8'],
                  ['s5', 's3'], ['s5', 's3'], ['s3', 's4', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'], ['s2', 's6'], ['s1', 's4'], ['s7'], ['s5', 's7'],
                  ['s4'], ['s1', 's8'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7'], ['s4', 's7'], ['s6'], ['s5'], ['s3'], ['s1', 's2'],
                  ['s3', 's6'], ['s1', 's2'], ['s4', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8'], ['s8', 's1'], ['s3', 's6'], ['s2'], ['s3', 's6'], ['s1'], ['s8'], ['s2', 's8'],
                  ['s2'], ['s4', 's3'], ['s3']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10'], ['s6'], ['s7'], ['s3'], ['s1', 's7'], ['s5'], ['s3'], ['s6', 's10'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9'], ['s9'], ['s9', 's2'], ['s10'], ['s6', 's8'], ['s6'], ['s2'], ['s5', 's10'], ['s2'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6'], ['s8'], ['s4'], ['s10'], ['s7', 's10'], ['s6'], ['s5', 's4'], ['s8'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3'], ['s3'], ['s4', 's9'], ['s7'], ['s1'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s5', 's10'], ['s10'], ['s8', 's4'], ['s8'], ['s4'], ['s4', 's7'], ['s1', 's3'], ['s10', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3'], ['s1', 's4'], ['s5'], ['s3'], ['s5'], ['s5', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s7'], ['s2'], ['s5'], ['s2', 's10'], ['s5'], ['s8'], ['s7'], ['s10'], ['s7'], ['s5', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s3', 's5'], ['s2'], ['s10'], ['s4', 's6'], ['s3'], ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10'], ['s10'], ['s1', 's2'], ['s5'], ['s3'],
                  ['s2', 's6'], ['s1', 's2', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'], ['s2', 's9'], ['s8'], ['s5'], ['s2'], ['s1'], ['s7']])])

    elif vehicle_number == 16:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'], ['s7'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6'], ['s1', 's7'], ['s2', 's6'], ['s1', 's3'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'], ['s5'], ['s1', 's6'], ['s5'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s6'], ['s8', 's5'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s1'], ['s1'], ['s8'],
                  ['s2', 's4'], ['s7', 's2'], ['s1', 's2', 's7'], ['s1', 's4'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s7'], ['s1'], ['s4', 's7'], ['s1', 's2'], ['s1'], ['s4', 's2'], ['s2', 's7'],
                  ['s2'], ['s4'], ['s7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's3'], ['s3', 's2'], ['s1', 's2'], ['s2', 's8'], ['s8'], ['s7'], ['s1', 's8'], ['s1'], ['s2', 's3'], ['s4', 's7'], ['s1', 's7'], ['s7', 's4'], ['s7'],
                  ['s4'], ['s8', 's4'], ['s3', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s2', 's3'], ['s4', 's2'], ['s4'], ['s8'], ['s3'], ['s1', 's4'], ['s3'],
                  ['s1', 's2'], ['s1', 's3'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2'], ['s5'], ['s8'], ['s1'], ['s3'], ['s2', 's8'], ['s1', 's6'], ['s5', 's8'], ['s8'], ['s2', 's6'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s1'], ['s2'], ['s6'], ['s5'], ['s7'], ['s2', 's5'], ['s5'], ['s6'], ['s6', 's7'], ['s2'], ['s6', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s1', 's8'], ['s4'], ['s2', 's5'],['s8', 's5'], ['s8'], ['s5'], ['s1'], ['s5'], ['s1'], ['s5', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's6'], ['s6', 's9'], ['s1'], ['s7', 's4'], ['s4', 's2'], ['s2'], ['s2'], ['s6', 's9'],
                  ['s9'], ['s2', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s9', 's2'],
                  ['s7', 's9'], ['s8'], ['s2', 's1'], ['s4'], ['s4', 's8'], ['s2'], ['s7', 's2'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s6', 's2'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s9'], ['s2'], ['s8', 's2'], ['s3', 's7'],
                  ['s6'], ['s7'], ['s6', 's8'], ['s9'], ['s7'], ['s8', 's2'], ['s7'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s3', 's4'], ['s2'], ['s1', 's7'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'], ['s3'], ['s3', 's4'], ['s1', 's7'], ['s5'],
                  ['s5', 's2'], ['s7', 's6'], ['s2', 's5'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'], ['s2'], ['s6'], ['s5'], ['s1', 's4'],
                  ['s3', 's5'], ['s2', 's3'], ['s1'], ['s3', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'], ['s7'], ['s1', 's4'], ['s2', 's9'],
                  ['s2'], ['s8', 's9'], ['s1', 's4'], ['s7'], ['s8', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'], ['s8'], ['s1', 's7'], ['s2'], ['s7'],
                  ['s2', 's4'], ['s2'], ['s4', 's8'], ['s1', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'], ['s3', 's4'], ['s1', 's8'], ['s7'], ['s8'],
                  ['s5', 's3'], ['s5', 's3'], ['s3', 's4', 's5'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'], ['s2', 's6'], ['s1', 's4'], ['s7'], ['s5', 's7'],
                  ['s4'], ['s1', 's8'], ['s8', 's4'], ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7', 's5'], ['s4', 's7'], ['s6'], ['s5'], ['s3'], ['s1', 's2'],
                  ['s3', 's6'], ['s1', 's2'], ['s4', 's7'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8', 's1'], ['s8', 's1'], ['s3', 's6'], ['s2'], ['s3', 's6'], ['s1'], ['s8'], ['s2', 's8'],
                  ['s2'], ['s4', 's3'], ['s3'], ['s4', 's6']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10', 's6'], ['s6', 's10'], ['s7'], ['s5'],
                  ['s3'], ['s1', 's7'], ['s5'], ['s3'], ['s6', 's10'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9', 's2'], ['s9'], ['s9', 's2'], ['s10'],
                  ['s6', 's8'], ['s6'], ['s2'], ['s5', 's10'], ['s2', 's9'], ['s1'], ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6', 's8'], ['s8'], ['s4', 's5'], ['s4'],
                  ['s10'], ['s7', 's10'], ['s6'], ['s5', 's4'], ['s8'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3', 's7'],
                  ['s3'], ['s4', 's9'], ['s7'], ['s1'], ['s5'], ['s1', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s1', 's3'], ['s5', 's10'], ['s10'],
                  ['s8', 's4'], ['s8'], ['s4'], ['s4', 's7'], ['s1', 's3'], ['s10', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3'],
                  ['s1', 's4'], ['s5'], ['s3'], ['s5'], ['s5', 's7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s2', 's10'], ['s7'], ['s2'], ['s5'],
                  ['s2', 's10'], ['s5'], ['s8'], ['s7'], ['s10'], ['s7'], ['s5', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'],
                  ['s3', 's5'], ['s2'], ['s10'], ['s4', 's6'], ['s3'], ['s1', 's10'], ['s3', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10', 's4'], ['s10'],
                  ['s1', 's2'], ['s5'], ['s3'], ['s3', 's5'], ['s2', 's6'], ['s1', 's2', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9', 's2'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'],
                  ['s2', 's9'], ['s8'], ['s5'], ['s2'], ['s1'], ['s7'], ['s7', 's1']])])

    elif vehicle_number == 17:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'], ['s7'], ['s1'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6'], ['s1', 's7'], ['s2', 's6'], ['s1', 's3'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'], ['s5'], ['s1', 's6'], ['s5'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s6'], ['s8', 's5'], ['s1', 's6'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s8', 's4'], ['s8', 's2'], ['s7', 's4'], ['s7', 's8'], ['s7'], ['s4'], ['s8'], ['s2'], ['s1'], ['s1'], ['s8'],
                  ['s2', 's4'], ['s7', 's2'], ['s1', 's2', 's7'], ['s1', 's4'], ['s1'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s4', 's1'], ['s2', 's4'], ['s2', 's1'], ['s7', 's4'], ['s7'], ['s7'], ['s1'], ['s4', 's7'], ['s1', 's2'], ['s1'], ['s4', 's2'], ['s2', 's7'],
                  ['s2'], ['s4'], ['s7'], ['s1'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's3'], ['s3', 's2'], ['s1', 's2'], ['s2', 's8'], ['s8'], ['s7'], ['s1', 's8'], ['s1'], ['s2', 's3'], ['s4', 's7'], ['s1', 's7'], ['s7', 's4'], ['s7'],
                  ['s4'], ['s8', 's4'], ['s3', 's8'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s2', 's3'], ['s4', 's2'], ['s4'], ['s8'], ['s3'], ['s1', 's4'], ['s3'],
                  ['s1', 's2'], ['s1', 's3'], ['s3'], ['s3'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2', 's6'], ['s2'], ['s5'], ['s8'], ['s1'], ['s3'], ['s2', 's8'], ['s1', 's6'], ['s5', 's8'], ['s8'], ['s2', 's6'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s2', 's6'], ['s1', 's7'], ['s5', 's1'], ['s7', 's5'], ['s2', 's6'], ['s1'], ['s2'], ['s6'], ['s5'], ['s7'], ['s2', 's5'], ['s5'], ['s6'], ['s6', 's7'], ['s2'], ['s6', 's1'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s4', 's5'], ['s1', 's4'], ['s8', 's1'], ['s8', 's4'], ['s8'], ['s1', 's8'], ['s1', 's8'], ['s4'], ['s2', 's5'],['s8', 's5'], ['s8'], ['s5'], ['s1'], ['s5'], ['s1'], ['s5', 's4'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's4'], ['s1', 's6'], ['s6', 's9'], ['s1'], ['s7', 's4'], ['s4', 's2'], ['s2'], ['s2'], ['s6', 's9'],
                  ['s9'], ['s2', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s9', 's2'], ['s4', 's8'],
                  ['s7', 's9'], ['s8'], ['s2', 's1'], ['s4'], ['s4', 's8'], ['s2'], ['s7', 's2'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s6', 's2'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s9'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s8', 's2'],
                  ['s6'], ['s7'], ['s6', 's8'], ['s9'], ['s7'], ['s8', 's2'], ['s7'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s3', 's4'], ['s2'], ['s1', 's7'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'], ['s6', 's7'], ['s3'], ['s3', 's4'], ['s1', 's7'], ['s5'],
                  ['s5', 's2'], ['s7', 's6'], ['s2', 's5'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'], ['s2'], ['s6'], ['s1', 's4'], ['s5'], ['s1', 's4'],
                  ['s3', 's5'], ['s2', 's3'], ['s1'], ['s3', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'], ['s7'], ['s1', 's4'], ['s2', 's8'], ['s2', 's9'],
                  ['s2'], ['s8', 's9'], ['s1', 's4'], ['s7'], ['s8', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s4', 's8'], ['s7'], ['s1'], ['s4'], ['s8', 's2'], ['s7', 's3', 's1'], ['s3'], ['s1', 's3'], ['s8'], ['s1', 's7'], ['s2'], ['s7'], ['s2', 's8'],
                  ['s2', 's4'], ['s2'], ['s4', 's8'], ['s1', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s4', 's5'], ['s4', 's3'], ['s7', 's8'], ['s1'], ['s7'], ['s8', 's1'], ['s4'], ['s1', 's7'], ['s3', 's4'], ['s1', 's8'], ['s7'], ['s8'], ['s1'],
                  ['s5', 's3'], ['s5', 's3'], ['s3', 's4', 's5'], ['s5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'], ['s2', 's6'], ['s1', 's4'], ['s7'], ['s5', 's7'], ['s8', 's4'],
                  ['s4'], ['s1', 's8'], ['s8', 's4'], ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7', 's5'], ['s4', 's7'], ['s6'], ['s5'], ['s3'], ['s1', 's2'], ['s4', 's5'],
                  ['s3', 's6'], ['s1', 's2'], ['s4', 's7'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s3', 's6'], ['s4'], ['s4', 's6'], ['s1', 's2'], ['s8', 's1'], ['s8', 's1'], ['s3', 's6'], ['s2'], ['s3', 's6'], ['s1'], ['s8'], ['s2', 's8'], ['s1', 's2'],
                  ['s2'], ['s4', 's3'], ['s3'], ['s4', 's6']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10', 's6'], ['s6', 's10'], ['s7'], ['s5'], ['s5'],
                  ['s3'], ['s1', 's7'], ['s5'], ['s3'], ['s6', 's10'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's5'], ['s6', 's8'], ['s5', 's10'], ['s1'], ['s8'], ['s9', 's2'], ['s9'], ['s9', 's2'], ['s10'], ['s6', 's8'],
                  ['s6', 's8'], ['s6'], ['s2'], ['s5', 's10'], ['s2', 's9'], ['s1'], ['s1', 's10']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4', 's5'], ['s5'], ['s6', 's8'], ['s8'], ['s4', 's5'], ['s4'], ['s6'],
                  ['s10'], ['s7', 's10'], ['s6'], ['s5', 's4'], ['s8'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3', 's7'], ['s4', 's9'],
                  ['s3'], ['s4', 's9'], ['s7'], ['s1'], ['s5'], ['s1', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s7'], ['s3'], ['s1', 's3'], ['s5', 's10'], ['s10'], ['s8', 's7', 's4'],
                  ['s8', 's4'], ['s8'], ['s4'], ['s4', 's7'], ['s1', 's3'], ['s10', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s7', 's5'], ['s1'], ['s3'], ['s3'],
                  ['s1', 's4'], ['s5'], ['s3'], ['s5'], ['s5', 's7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s5', 's8'], ['s2', 's10'], ['s10'], ['s8'], ['s7'], ['s2', 's10'], ['s7'], ['s2'], ['s5'], ['s8'],
                  ['s2', 's10'], ['s5'], ['s8'], ['s7'], ['s10'], ['s7'], ['s5', 's8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s6', 's4'],
                  ['s3', 's5'], ['s2'], ['s10'], ['s4', 's6'], ['s3'], ['s1', 's10'], ['s3', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10', 's4'], ['s10'], ['s10'],
                  ['s1', 's2'], ['s5'], ['s3'], ['s3', 's5'], ['s2', 's6'], ['s1', 's2', 's6'], ['s4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17'],
                 [['s7'], ['s7', 's1'], ['s8'], ['s9', 's2'], ['s9'], ['s5', 's8'], ['s9', 's2'], ['s1'], ['s5'], ['s8'],
                  ['s2', 's9'], ['s8'], ['s5'], ['s2'], ['s1'], ['s7'], ['s7', 's1']])])

    elif vehicle_number == 18:
        if subfleet_number == 1:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's7'], ['s3', 's1'], ['s3', 's2'], ['s2', 's6'], ['s6'], ['s7'], ['s2'], ['s7'], ['s1'], ['s3'],
                  ['s1', 's3'], ['s3', 's7'], ['s6', 's7'], ['s2', 's6'], ['s1', 's7'], ['s2', 's6'], ['s1', 's3'], ['s6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's6'], ['s8', 's1'], ['s5', 's6'], ['s5', 's1'], ['s5'], ['s1'], ['s5'], ['s5'], ['s1', 's6'], ['s5'], ['s8'],
                  ['s8', 's6'], ['s8'], ['s8', 's6'], ['s6'], ['s8', 's5'], ['s1', 's6'], ['s8']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's3'], ['s3', 's2'], ['s1', 's2'], ['s2', 's8'], ['s8'], ['s7'], ['s1', 's8'], ['s1'], ['s2', 's3'], ['s4', 's7'], ['s1', 's7'], ['s7', 's4'], ['s7'],
                  ['s4'], ['s8', 's4'], ['s3', 's8'], ['s3'], ['s2', 's4']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s8', 's2'], ['s1', 's4'], ['s1', 's8'], ['s8', 's4'], ['s2'], ['s2', 's3'], ['s4', 's2'], ['s4'], ['s8'], ['s3'], ['s1', 's4'], ['s3'],
                  ['s1', 's2'], ['s1', 's3'], ['s3'], ['s3'], ['s4'], ['s3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's3'], ['s2', 's1'], ['s6', 's5'], ['s2', 's6'], ['s2'], ['s5'], ['s8'], ['s1'], ['s3'],
                  ['s2', 's8'], ['s1', 's6'], ['s5', 's8'], ['s2'], ['s8'], ['s2', 's6'], ['s1', 's3'], ['s5'], ['s6', 's3']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s7', 's4'], ['s1', 's4'], ['s6', 's4'], ['s7'], ['s9', 's7'], ['s1'], ['s1', 's4'],
                  ['s1', 's6'], ['s6', 's9'], ['s1'], ['s7', 's4'], ['s4', 's2'], ['s2'], ['s2'], ['s1', 's6'], ['s6', 's9'],
                  ['s9'], ['s2', 's7']])])
        # con 2 subflotas
        elif subfleet_number == 2:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s9', 's7'], ['s1'], ['s4', 's8'], ['s9', 's1'], ['s8'], ['s1', 's7'], ['s4'], ['s9', 's2'], ['s4', 's8'],
                  ['s7', 's9'], ['s2'], ['s8'], ['s2', 's1'], ['s4'], ['s4', 's8'], ['s2'], ['s7', 's2'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s6', 's2'], ['s9', 's3'], ['s3'], ['s6', 's8'], ['s9'], ['s2'], ['s8', 's2'], ['s3', 's7'], ['s8', 's2'],
                  ['s6'], ['s7'], ['s2'], ['s6', 's8'], ['s9'], ['s7'], ['s8', 's2'], ['s7'], ['s3', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s3', 's4'], ['s2'], ['s1', 's7'], ['s6', 's7'], ['s4', 's2'], ['s5', 's4'], ['s1', 's6'], ['s3'], ['s6', 's7'], ['s3'], ['s3', 's4'], ['s1', 's7'], ['s5'],
                  ['s5', 's2'], ['s2', 's3'], ['s7', 's6'], ['s2', 's5'], ['s1', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's4'], ['s3', 's2'], ['s5'], ['s4'], ['s6'], ['s5', 's2'], ['s6', 's1'], ['s4', 's6'], ['s2'], ['s6'], ['s1', 's4'], ['s5'], ['s1', 's4'],
                  ['s3', 's5'], ['s3', 's5'], ['s2', 's3'], ['s1'], ['s3', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s4'], ['s7', 's1'], ['s7', 's4'], ['s1'], ['s2', 's8'], ['s8', 's9'], ['s2', 's9'], ['s1', 's4'], ['s7'], ['s1', 's4'], ['s2', 's8'], ['s2', 's9'],
                  ['s2'], ['s8', 's9'], ['s7'], ['s1', 's4'], ['s7'], ['s8', 's9']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's8'], ['s4', 's8'], ['s1'], ['s5', 's7', 's2'], ['s6', 's5'], ['s6'], ['s7'], ['s2', 's5'], ['s2', 's6'], ['s1', 's4'], ['s7'], ['s5', 's7'], ['s8', 's4'],
                  ['s4'], ['s1'], ['s1', 's8'], ['s8', 's4'], ['s2', 's6']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's2', 's6'], ['s5', 's4'], ['s3', 's1'], ['s2'], ['s5', 's7'], ['s4'], ['s7', 's5'], ['s4', 's7'], ['s6'], ['s5'], ['s3'], ['s1', 's2'], ['s4', 's5'],
                  ['s3', 's6'], ['s3'], ['s1', 's2'], ['s4', 's7'], ['s6', 's3']])])
        # con 3 subflotas
        elif subfleet_number == 3:
            return random.choice([
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s3', 's1'], ['s6', 's10'], ['s3', 's7'], ['s1'], ['s5'], ['s5'], ['s10', 's6'], ['s6', 's10'], ['s7'], ['s5'], ['s5'],
                  ['s3'], ['s1', 's7'], ['s5'], ['s3'], ['s6'], ['s6', 's10'], ['s1', 's7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s7', 's10'], ['s6', 's8'], ['s7'], ['s4'], ['s4', 's5'], ['s5'], ['s6', 's8'], ['s8'], ['s4', 's5'], ['s4'], ['s6'],
                  ['s10'], ['s7', 's10'], ['s6'], ['s5', 's4'], ['s8'], ['s10'], ['s7']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's5'], ['s4', 's9'], ['s7'], ['s3'], ['s4', 's9'], ['s3', 's7'], ['s1'], ['s4'], ['s9'], ['s5'], ['s3', 's7'], ['s4', 's9'],
                  ['s3'], ['s4', 's9'], ['s7'], ['s1'], ['s5'], ['s1', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's3'], ['s5'], ['s1'], ['s5'], ['s8', 's7'], ['s10'], ['s7'], ['s3'], ['s1', 's3'], ['s5', 's10'], ['s10'], ['s8', 's7', 's4'],
                  ['s8', 's4'], ['s8'], ['s4'], ['s4', 's7'], ['s1', 's3'], ['s10', 's5']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s7'], ['s4'], ['s3'], ['s4'], ['s7'], ['s3'], ['s4', 's1'], ['s1'], ['s7', 's5'], ['s1'], ['s3'], ['s3'],
                  ['s1', 's4'], ['s5'], ['s3'], ['s5'], ['s5', 's7'], ['s1']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's2'], ['s2', 's10'], ['s6', 's4'], ['s3', 's5'], ['s10'], ['s4'], ['s6'], ['s5'], ['s1'], ['s6', 's4'],
                  ['s3', 's5'], ['s2'], ['s10'], ['s4', 's6'], ['s3'], ['s1', 's10'], ['s3', 's5'], ['s1', 's2']]),
                (['v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18'],
                 [['s1', 's2'], ['s3', 's5'], ['s1', 's6'], ['s4'], ['s5'], ['s4', 's10'], ['s3'], ['s10', 's4'], ['s10'], ['s10'],
                  ['s1', 's2'], ['s5'], ['s3'], ['s3', 's5'], ['s2', 's6'], ['s1', 's2', 's6'], ['s4'], ['s6']])])

# if __name__ == '__main__':
#     import networkx as nx
#     import matplotlib.pyplot as plt
#
#     cantidad_de_subflotas = 4
#
#
#     def test_configuration(vs, ss):
#         def obtain_subfleets():
#             ss_sf = []
#             cant = 0
#             i = 0
#             for node_p in set_p.nodes(data=True):
#                 j = 0
#                 for node_q in set_p.nodes(data=True):
#                     j += 1
#                     if i < j:
#                         intersection = sorted(node_p[1]["S_p"].keys() & node_q[1]["S_p"].keys())
#                         if node_p != node_q and len(intersection) > 0:
#                             if not set_p.has_edge(node_p[0], node_q[0]):
#                                 set_p.add_edge(node_p[0], node_q[0], S_pq=intersection)
#                 i += 1
#             sub = sorted(nx.connected_components(set_p))
#             sub_fleets = [sorted(item) for item in sub]
#             for i, sub_fleet in enumerate(sub_fleets):
#                 sub_fleet = sorted(sub_fleet)
#                 s_sf = set()
#                 for particle in sub_fleet:
#                     s_sf = s_sf | set_p.nodes[particle]['S_p'].keys()
#                 s_sf = sorted(s_sf)
#                 for j, sensor in enumerate(s_sf):
#                     list_vehicles = list()
#                     for particle in sub_fleet:
#                         v_sensors = set_p.nodes[particle]['S_p'].keys()
#                         for key in v_sensors:
#                             if key == sensor:
#                                 list_vehicles.append(particle)
#                 ss_sf.append(s_sf)
#
#                 for particle in sub_fleet:
#                     sensors = set_p.nodes[particle]['S_p'].keys()
#                     for s, sensor in enumerate(sensors):
#                         cant += 1
#             return sub_fleets, ss_sf, cant
#
#         set_p = nx.MultiGraph()
#         for p, (part, sen) in enumerate(zip(vs, ss)):
#             set_p.add_node(part, S_p=dict.fromkeys(sen, []), index=p, )
#
#         return set_p, obtain_subfleets()
#
#
#     vehicles, sensores_in_vehicles = obtain_prefabricated_vehicles(cantidad_de_subflotas)
#
#     p, (vsf, ssf, cant) = test_configuration(vehicles, sensores_in_vehicles)
#
#     print(f'{vehicles=}')
#     print(f'{sensores_in_vehicles=}')
#     print(f'numero_total_de_sensores={cant}')
#
#     sub_fleets = nx.connected_components(p)
#     # Graficamos, para eso necesitamos diferentes colores para diferentes sf
#     colors = ["gold",
#               "red",
#               "limegreen",
#               "darkorange",
#               ]
#
#     # recorremos las subflotas
#     for i, sub_fleet in enumerate(sub_fleets):
#         # SOLO PARA GRAFICAR creamos nuevos grafos! SOLO PARA GRAFICAR !!!!!!
#         P_sf = p.subgraph(sub_fleet)
#
#         all_S_p = nx.get_node_attributes(P_sf, 'S_p')
#         labels = dict()
#         for vehicle in all_S_p:
#             labels[vehicle] = f"{vehicle} \n S({vehicle}): {list(all_S_p[vehicle].keys())}"
#         # como estamos creando grafos en cada iteracion, desplazamos los centros de las subflotas con
#         # center = []
#         pos_vehicles = nx.circular_layout(P_sf, center=[i * 2.5, 0])
#         axis = plt.gca()
#         # graficamos y luego además agregamos el color de los nodos con node_color
#         nx.draw(P_sf, pos_vehicles, labels=labels, ax=axis, node_color=colors[i], node_size=5000)
#         # graficamos los edges
#         edge_labels = dict([((u, v,), d['S_pq'])
#                             for u, v, d in P_sf.edges(data=True)])
#         nx.draw_networkx_edge_labels(P_sf, pos_vehicles, edge_labels=edge_labels, label_pos=0.3, font_size=15)
#         plt.tight_layout()
#     plt.show()
